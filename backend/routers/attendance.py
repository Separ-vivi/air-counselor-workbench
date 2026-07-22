"""查课考勤管理 API"""
import logging
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, desc, or_, and_
from sqlalchemy.orm import Session
from database import get_db
from models import Student, StudentAttendanceException, ClassModel
from pydantic import BaseModel
from typing import Optional

logger = logging.getLogger(__name__)

router = APIRouter(prefix='/api/attendance', tags=['查课考勤'])


# ===== 学期日期映射 =====

def _semester_date_range(semester: str):
    """将学期字符串转为日期范围"""
    if not semester or semester == 'all':
        return None, None
    parts = semester.split('-')
    if len(parts) == 3:
        y1, y2, term = parts[0], parts[1], parts[2]
        if term == '1':
            return f"{y1}-09-01", f"{y2}-01-31"
        else:
            return f"{y2}-02-01", f"{y2}-07-31"
    return None, None


# ===== Pydantic 模型 =====

class AttendanceCreate(BaseModel):
    student_id: int
    exception_date: str
    course_name: str = ''
    exception_type: str = '旷课'
    notes: str = ''


class AttendanceUpdate(BaseModel):
    exception_date: Optional[str] = None
    course_name: Optional[str] = None
    exception_type: Optional[str] = None
    notes: Optional[str] = None


# ===== CRUD =====

@router.get('/')
def list_attendance(
    class_id: int = Query(None),
    exception_type: str = Query(None),
    keyword: str = Query(None),
    semester: str = Query(None),
    date_from: str = Query(None),
    date_to: str = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    """获取考勤异常列表"""
    query = db.query(StudentAttendanceException).join(Student)

    # 班级筛选
    if class_id:
        query = query.filter(Student.class_id == class_id)

    # 异常类型筛选
    if exception_type:
        query = query.filter(StudentAttendanceException.exception_type == exception_type)

    # 关键词搜索
    if keyword:
        query = query.filter(or_(
            Student.name.contains(keyword),
            Student.student_no.contains(keyword),
            StudentAttendanceException.course_name.contains(keyword)
        ))

    # 学期筛选（映射日期范围）
    if semester:
        s_start, s_end = _semester_date_range(semester)
        if s_start and s_end:
            query = query.filter(
                StudentAttendanceException.exception_date >= s_start,
                StudentAttendanceException.exception_date <= s_end
            )

    # 日期范围筛选
    if date_from:
        query = query.filter(StudentAttendanceException.exception_date >= date_from)
    if date_to:
        query = query.filter(StudentAttendanceException.exception_date <= date_to)

    total = query.count()
    items = query.order_by(desc(StudentAttendanceException.exception_date)) \
        .offset((page - 1) * size).limit(size).all()

    result = []
    for item in items:
        student = item.student
        result.append({
            'id': item.id,
            'student_id': item.student_id,
            'student_no': student.student_no,
            'student_name': student.name,
            'class_name': student.class_name,
            'exception_date': item.exception_date,
            'course_name': item.course_name,
            'exception_type': item.exception_type,
            'notes': item.notes,
            'created_at': item.created_at.isoformat() if item.created_at else None
        })

    return {'total': total, 'items': result}


@router.post('/')
def create_attendance(data: AttendanceCreate, db: Session = Depends(get_db)):
    """新增考勤异常记录"""
    item = StudentAttendanceException(
        student_id=data.student_id,
        exception_date=data.exception_date,
        course_name=data.course_name,
        exception_type=data.exception_type,
        notes=data.notes
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {'id': item.id, 'message': '创建成功'}


@router.put('/{item_id}')
def update_attendance(item_id: int, data: AttendanceUpdate, db: Session = Depends(get_db)):
    """编辑考勤异常记录"""
    item = db.query(StudentAttendanceException).filter(StudentAttendanceException.id == item_id).first()
    if not item:
        raise HTTPException(404, '记录不存在')

    if data.exception_date is not None:
        item.exception_date = data.exception_date
    if data.course_name is not None:
        item.course_name = data.course_name
    if data.exception_type is not None:
        item.exception_type = data.exception_type
    if data.notes is not None:
        item.notes = data.notes

    db.commit()
    return {'message': '更新成功'}


@router.delete('/{item_id}')
def delete_attendance(item_id: int, db: Session = Depends(get_db)):
    """删除考勤异常记录"""
    item = db.query(StudentAttendanceException).filter(StudentAttendanceException.id == item_id).first()
    if not item:
        raise HTTPException(404, '记录不存在')

    db.delete(item)
    db.commit()
    return {'message': '删除成功'}


# ===== 统计 API =====

@router.get('/stats')
def get_stats(db: Session = Depends(get_db)):
    """统计概览：总异常数、涉及学生数、按类型分布、本周新增、本月新增"""
    today = datetime.now().date()
    # 本周（周一到周日）
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    # 本月
    month_start = today.replace(day=1)

    week_start_str = week_start.strftime('%Y-%m-%d')
    week_end_str = week_end.strftime('%Y-%m-%d')
    month_start_str = month_start.strftime('%Y-%m-%d')
    today_str = today.strftime('%Y-%m-%d')

    total = db.query(StudentAttendanceException).count()
    student_count = db.query(
        func.count(func.distinct(StudentAttendanceException.student_id))
    ).scalar() or 0

    # 按类型分布
    type_rows = db.query(
        StudentAttendanceException.exception_type,
        func.count(StudentAttendanceException.id)
    ).group_by(StudentAttendanceException.exception_type).all()
    by_type = {t or '未知': c for t, c in type_rows}

    # 本周新增
    week_new = db.query(StudentAttendanceException).filter(
        StudentAttendanceException.exception_date >= week_start_str,
        StudentAttendanceException.exception_date <= week_end_str
    ).count()

    # 本月新增
    month_new = db.query(StudentAttendanceException).filter(
        StudentAttendanceException.exception_date >= month_start_str,
        StudentAttendanceException.exception_date <= today_str
    ).count()

    # 旷课占比
    absent_count = by_type.get('旷课', 0)
    absent_ratio = round(absent_count / total * 100, 1) if total > 0 else 0

    return {
        'total': total,
        'student_count': student_count,
        'by_type': by_type,
        'week_new': week_new,
        'month_new': month_new,
        'absent_ratio': absent_ratio
    }


@router.get('/top-students')
def get_top_students(
    limit: int = Query(10, ge=1, le=50),
    db: Session = Depends(get_db)
):
    """高频异常学生 TOP N（按异常次数降序）"""
    rows = db.query(
        StudentAttendanceException.student_id,
        func.count(StudentAttendanceException.id).label('cnt')
    ).group_by(StudentAttendanceException.student_id) \
     .order_by(desc('cnt')) \
     .limit(limit).all()

    result = []
    for student_id, cnt in rows:
        student = db.query(Student).filter(Student.id == student_id).first()
        if student:
            result.append({
                'student_id': student.id,
                'student_name': student.name,
                'student_no': student.student_no,
                'class_name': student.class_name,
                'count': cnt
            })
    return result


@router.get('/top-courses')
def get_top_courses(
    limit: int = Query(10, ge=1, le=50),
    db: Session = Depends(get_db)
):
    """高频异常课程 TOP N"""
    rows = db.query(
        StudentAttendanceException.course_name,
        func.count(StudentAttendanceException.id).label('cnt')
    ).group_by(StudentAttendanceException.course_name) \
     .order_by(desc('cnt')) \
     .limit(limit).all()

    return [{'course_name': name or '未知', 'count': cnt} for name, cnt in rows]


@router.get('/monthly-trend')
def get_monthly_trend(db: Session = Depends(get_db)):
    """月度趋势：最近6个月每月异常次数"""
    today = datetime.now().date()
    months = []
    for i in range(5, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        months.append((y, m))

    result = []
    for y, m in months:
        start = f"{y}-{m:02d}-01"
        if m == 12:
            end = f"{y + 1}-01-01"
        else:
            end = f"{y}-{m + 1:02d}-01"
        count = db.query(StudentAttendanceException).filter(
            StudentAttendanceException.exception_date >= start,
            StudentAttendanceException.exception_date < end
        ).count()
        result.append({'month': f"{y}-{m:02d}", 'count': count})

    return result


# ===== 导出 Excel =====

@router.get('/export')
def export_attendance(
    class_id: int = Query(None),
    exception_type: str = Query(None),
    keyword: str = Query(None),
    semester: str = Query(None),
    date_from: str = Query(None),
    date_to: str = Query(None),
    db: Session = Depends(get_db)
):
    """导出考勤异常 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(500, 'openpyxl 未安装')

    from io import BytesIO

    # 复用列表查询逻辑（不分页，取全部）
    query = db.query(StudentAttendanceException).join(Student)
    if class_id:
        query = query.filter(Student.class_id == class_id)
    if exception_type:
        query = query.filter(StudentAttendanceException.exception_type == exception_type)
    if keyword:
        query = query.filter(or_(
            Student.name.contains(keyword),
            Student.student_no.contains(keyword),
            StudentAttendanceException.course_name.contains(keyword)
        ))
    if semester:
        s_start, s_end = _semester_date_range(semester)
        if s_start and s_end:
            query = query.filter(
                StudentAttendanceException.exception_date >= s_start,
                StudentAttendanceException.exception_date <= s_end
            )
    if date_from:
        query = query.filter(StudentAttendanceException.exception_date >= date_from)
    if date_to:
        query = query.filter(StudentAttendanceException.exception_date <= date_to)

    items = query.order_by(desc(StudentAttendanceException.exception_date)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '考勤异常'

    headers = ['学号', '姓名', '班级', '日期', '课程', '异常类型', '备注']
    header_fill = PatternFill(start_color='5B92E5', end_color='5B92E5', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(items, 2):
        student = item.student
        ws.cell(row=row_idx, column=1, value=student.student_no)
        ws.cell(row=row_idx, column=2, value=student.name)
        ws.cell(row=row_idx, column=3, value=student.class_name)
        ws.cell(row=row_idx, column=4, value=item.exception_date)
        ws.cell(row=row_idx, column=5, value=item.course_name)
        ws.cell(row=row_idx, column=6, value=item.exception_type)
        ws.cell(row=row_idx, column=7, value=item.notes or '')

    # 列宽自适应
    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, len(items) + 2))
        ws.column_dimensions[chr(64 + col)].width = min(max_len + 4, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=attendance_exceptions.xlsx'}
    )
