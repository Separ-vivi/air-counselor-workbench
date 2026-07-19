"""学生干部 + 班主任 + 就业升学 路由"""
from fastapi import APIRouter, Depends, HTTPException, Query, Body
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional
from database import get_db
from models import StudentCadreRecord, ClassTeacher, EmploymentRecord, Student, ClassModel

router = APIRouter(prefix='/api')


def _xlsx_stream(wb, filename):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


def _get_student_class_name(db: Session, student_id: int) -> str:
    """通过student_id获取班级名称"""
    stu = db.query(Student).filter(Student.id == student_id).first()
    if stu and stu.class_obj:
        return stu.class_obj.class_name
    return ''


# ===== 学生干部 =====
def _cadre_dict(c, db):
    stu = db.query(Student).filter(Student.id == c.student_id).first()
    return {
        'id': c.id, 'student_id': c.student_id,
        'class_name': _get_student_class_name(db, c.student_id),
        'student_name': stu.name if stu else '',
        'student_no': stu.student_no if stu else '',
        'phone': stu.phone if stu else '',
        'position': c.position,
        'level': getattr(c, 'level', '') or '',
        'organization': getattr(c, 'organization', '') or '',
        'term': c.term,
        'start_date': getattr(c, 'start_date', '') or '',
        'end_date': getattr(c, 'end_date', '') or '',
        'email': getattr(c, 'email', '') or '',
        'notes': c.notes,
    }


@router.get('/cadres')
def list_cadres(
    class_name: Optional[str] = Query(None),
    level: Optional[str] = Query(None),
    position: Optional[str] = Query(None),
    student_id: Optional[int] = Query(None),
    search: str = Query('', description='搜索学号/姓名/职务/组织'),
    sort_by: str = Query('position', description='排序字段'),
    order: str = Query('asc', description='asc/desc'),
    db: Session = Depends(get_db)
):
    """学生干部列表 (v3j-B-b03 · 支持 search + sort_by + order + class_name 过滤)"""
    q = db.query(StudentCadreRecord).outerjoin(Student, StudentCadreRecord.student_id == Student.id)
    # class_name 过滤：在数据库层通过 join ClassModel 实现，支持模糊匹配
    if class_name:
        q = q.outerjoin(ClassModel, Student.class_id == ClassModel.id).filter(
            ClassModel.class_name.contains(class_name.strip())
        )
    if level:
        q = q.filter(StudentCadreRecord.level == level)
    if position:
        q = q.filter(StudentCadreRecord.position.contains(position))
    if student_id:
        q = q.filter(StudentCadreRecord.student_id == student_id)
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Student.name.ilike(pattern),
                Student.student_no.ilike(pattern),
                StudentCadreRecord.position.ilike(pattern),
                StudentCadreRecord.organization.ilike(pattern),
                StudentCadreRecord.level.ilike(pattern),
            )
        )
    SORT_WHITELIST = {
        'position': StudentCadreRecord.position,
        'level': StudentCadreRecord.level,
        'organization': StudentCadreRecord.organization,
        'start_date': StudentCadreRecord.start_date,
        'end_date': StudentCadreRecord.end_date,
        'student_name': Student.name,
        'student_no': Student.student_no,
    }
    col = SORT_WHITELIST.get(sort_by, StudentCadreRecord.position)
    if (order or 'asc').lower() == 'desc':
        q = q.order_by(col.desc())
    else:
        q = q.order_by(col.asc())
    items = q.all()
    return [_cadre_dict(c, db) for c in items]


def _cadre_export_workbook(items, db):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '学生干部'
    ws.append(['学号', '姓名', '班级', '职务', '级别', '组织', '任职起始', '任职结束', '任期', '邮箱', '备注'])
    for c in items:
        d = _cadre_dict(c, db)
        ws.append([
            d['student_no'], d['student_name'], d['class_name'],
            d['position'], d['level'], d['organization'],
            d['start_date'], d['end_date'], d.get('term', ''),
            d.get('email', ''), d.get('notes', ''),
        ])
    return wb


@router.post('/cadres/export')
def export_cadres_by_ids(payload: dict = Body(...), db: Session = Depends(get_db)):
    """按 ID 列表批量导出学生干部 Excel (v3j-B-b03)"""
    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请传入非空的 ids 列表')
    items = db.query(StudentCadreRecord).filter(StudentCadreRecord.id.in_(ids)).all()
    wb = _cadre_export_workbook(items, db)
    return _xlsx_stream(wb, 'cadres_selected.xlsx')


@router.get('/cadres/export/all')
def export_cadres_all(
    search: str = Query(''),
    level: Optional[str] = Query(None),
    position: Optional[str] = Query(None),
    student_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """按当前搜索条件导出全部学生干部 Excel (v3j-B-b03)"""
    q = db.query(StudentCadreRecord).outerjoin(Student, StudentCadreRecord.student_id == Student.id)
    if level:
        q = q.filter(StudentCadreRecord.level == level)
    if position:
        q = q.filter(StudentCadreRecord.position.contains(position))
    if student_id:
        q = q.filter(StudentCadreRecord.student_id == student_id)
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Student.name.ilike(pattern),
                Student.student_no.ilike(pattern),
                StudentCadreRecord.position.ilike(pattern),
                StudentCadreRecord.organization.ilike(pattern),
                StudentCadreRecord.level.ilike(pattern),
            )
        )
    items = q.order_by(StudentCadreRecord.position).all()
    wb = _cadre_export_workbook(items, db)
    return _xlsx_stream(wb, 'cadres_all.xlsx')


@router.get('/cadres/directory')
def cadre_directory(db: Session = Depends(get_db)):
    """班委通讯录 - 返回扁平列表，每行一条干部记录"""
    from models import ClassModel
    result = []
    # 直接查询所有干部记录，join 学生表和班级表
    cadres = db.query(StudentCadreRecord).join(Student, StudentCadreRecord.student_id == Student.id).join(ClassModel, Student.class_id == ClassModel.id).all()
    for c in cadres:
        stu = db.query(Student).filter(Student.id == c.student_id).first()
        cls = db.query(ClassModel).filter(ClassModel.id == stu.class_id).first() if stu else None
        result.append({
            'student_name': stu.name if stu else '',
            'student_no': stu.student_no if stu else '',
            'class_name': cls.class_name if cls else '',
            'position': c.position,
            'level': c.level or '',
        })
    return result



@router.get('/cadres/{cadre_id}')
def get_cadre(cadre_id: int, db: Session = Depends(get_db)):
    c = db.query(StudentCadreRecord).filter(StudentCadreRecord.id == cadre_id).first()
    if not c:
        raise HTTPException(404, '记录不存在')
    return _cadre_dict(c, db)


@router.post('/cadres')
def create_cadre(data: dict, db: Session = Depends(get_db)):
    c = StudentCadreRecord(**data)
    db.add(c)
    db.commit()
    db.refresh(c)
    return {'id': c.id}


@router.put('/cadres/{cid}')
def update_cadre(cid: int, data: dict, db: Session = Depends(get_db)):
    c = db.query(StudentCadreRecord).filter(StudentCadreRecord.id == cid).first()
    if not c:
        raise HTTPException(404, '记录不存在')
    for k, v in data.items():
        if hasattr(c, k):
            setattr(c, k, v)
    db.commit()
    return {'ok': True}


@router.delete('/cadres/{cid}')
def delete_cadre(cid: int, db: Session = Depends(get_db)):
    c = db.query(StudentCadreRecord).filter(StudentCadreRecord.id == cid).first()
    if c:
        db.delete(c)
        db.commit()
    return {'ok': True}


# ===== 班主任 =====
def _teacher_dict(t, class_map):
    cls_name = class_map.get(t.class_id, '') if t.class_id else ''
    return {
        'id': t.id, 'class_id': t.class_id,
        'class_name': cls_name or getattr(t, 'class_name', ''),
        'name': t.name,
        'staff_no': t.staff_no, 'teacher_no': t.staff_no,  # alias
        'department': t.department,
        'phone': t.phone, 'office': t.office,
        'research_direction': t.research_direction,
        'title': getattr(t, 'title', '') or '',
        'email': getattr(t, 'email', '') or '',
        'notes': t.notes,
    }


@router.get('/class-teachers')
def list_class_teachers(
    search: str = Query('', description='搜索姓名/工号/院系/电话/邮箱'),
    sort_by: str = Query('name', description='排序字段'),
    order: str = Query('asc', description='asc/desc'),
    db: Session = Depends(get_db)
):
    """任课教师/班主任列表 (v3j-B-b03 · 支持 search + sort_by + order)"""
    q = db.query(ClassTeacher)
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                ClassTeacher.name.ilike(pattern),
                ClassTeacher.staff_no.ilike(pattern),
                ClassTeacher.department.ilike(pattern),
                ClassTeacher.phone.ilike(pattern),
                ClassTeacher.email.ilike(pattern),
                ClassTeacher.title.ilike(pattern),
            )
        )
    SORT_WHITELIST = {
        'name': ClassTeacher.name,
        'teacher_no': ClassTeacher.staff_no,
        'staff_no': ClassTeacher.staff_no,
        'department': ClassTeacher.department,
        'title': ClassTeacher.title,
        'phone': ClassTeacher.phone,
        'email': ClassTeacher.email,
    }
    col = SORT_WHITELIST.get(sort_by, ClassTeacher.name)
    if (order or 'asc').lower() == 'desc':
        q = q.order_by(col.desc())
    else:
        q = q.order_by(col.asc())
    items = q.all()
    class_map = {c.id: c.class_name for c in db.query(ClassModel).all()}
    return [_teacher_dict(t, class_map) for t in items]


def _teachers_export_workbook(items, class_map):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '任课教师'
    ws.append(['姓名', '工号', '所带班级', '院系', '职称', '电话', '邮箱', '办公室', '研究方向', '备注'])
    for t in items:
        d = _teacher_dict(t, class_map)
        ws.append([
            d['name'], d['teacher_no'], d['class_name'],
            d['department'], d['title'], d['phone'], d['email'],
            d.get('office', ''), d.get('research_direction', ''), d.get('notes', ''),
        ])
    return wb


@router.post('/class-teachers/export')
def export_class_teachers_by_ids(payload: dict = Body(...), db: Session = Depends(get_db)):
    """按 ID 列表批量导出班主任 Excel (v3j-B-b03)"""
    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请传入非空的 ids 列表')
    items = db.query(ClassTeacher).filter(ClassTeacher.id.in_(ids)).all()
    class_map = {c.id: c.class_name for c in db.query(ClassModel).all()}
    wb = _teachers_export_workbook(items, class_map)
    return _xlsx_stream(wb, 'class_teachers_selected.xlsx')


@router.get('/class-teachers/export/all')
def export_class_teachers_all(
    search: str = Query(''),
    db: Session = Depends(get_db)
):
    """按当前搜索条件导出全部班主任 Excel (v3j-B-b03)"""
    q = db.query(ClassTeacher)
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                ClassTeacher.name.ilike(pattern),
                ClassTeacher.staff_no.ilike(pattern),
                ClassTeacher.department.ilike(pattern),
                ClassTeacher.phone.ilike(pattern),
                ClassTeacher.email.ilike(pattern),
                ClassTeacher.title.ilike(pattern),
            )
        )
    items = q.order_by(ClassTeacher.name).all()
    class_map = {c.id: c.class_name for c in db.query(ClassModel).all()}
    wb = _teachers_export_workbook(items, class_map)
    return _xlsx_stream(wb, 'class_teachers_all.xlsx')


def _teacher_normalize_input(data: dict) -> dict:
    d = dict(data)
    if 'teacher_no' in d and 'staff_no' not in d:
        d['staff_no'] = d.pop('teacher_no')
    allowed = {'class_id','name','staff_no','department','phone','office','research_direction','title','email','notes'}
    return {k: v for k, v in d.items() if k in allowed}


@router.post('/class-teachers')
def create_class_teacher(data: dict, db: Session = Depends(get_db)):
    t = ClassTeacher(**_teacher_normalize_input(data))
    db.add(t)
    db.commit()
    db.refresh(t)
    return {'id': t.id}


@router.put('/class-teachers/{tid}')
def update_class_teacher(tid: int, data: dict, db: Session = Depends(get_db)):
    t = db.query(ClassTeacher).filter(ClassTeacher.id == tid).first()
    if not t:
        raise HTTPException(404, '记录不存在')
    for k, v in _teacher_normalize_input(data).items():
        if hasattr(t, k):
            setattr(t, k, v)
    db.commit()
    return {'ok': True}


@router.delete('/class-teachers/{tid}')
def delete_class_teacher(tid: int, db: Session = Depends(get_db)):
    t = db.query(ClassTeacher).filter(ClassTeacher.id == tid).first()
    if t:
        db.delete(t)
        db.commit()
    return {'ok': True}


# ===== 就业升学 =====
def _emp_dict(e, db):
    stu = db.query(Student).filter(Student.id == e.student_id).first()
    # v3j-C c01 · 补出 grade_name 供前端"仅毕业年级"过滤
    grade_name = ''
    try:
        if stu and stu.class_obj and stu.class_obj.major and stu.class_obj.major.grade:
            grade_name = stu.class_obj.major.grade.grade_name or ''
    except Exception:
        grade_name = ''
    return {
        'id': e.id, 'student_id': e.student_id,
        'student_name': stu.name if stu else '',
        'student_no': stu.student_no if stu else '',
        'class_name': _get_student_class_name(db, e.student_id),
        'grade_name': grade_name,
        'intention_type': e.intention_type, 'target_industry': e.target_industry,
        'target_position': e.target_position, 'internship_company': e.internship_company,
        # 前端字段 alias
        'company': e.internship_company or e.target_industry or '',
        'position': e.target_position or '',
        'work_location': '',
        'sign_date': e.offer_date or '',
        'salary': e.salary_range or '',
        'status': e.status, 'offer_date': e.offer_date,
        'salary_range': e.salary_range, 'notes': e.notes,
    }


@router.get('/employment')
def list_employment(
    search: str = Query('', description='搜索学号/姓名/单位/岗位/状态'),
    status: Optional[str] = Query(None),
    student_id: Optional[int] = Query(None),
    sort_by: str = Query('id', description='排序字段'),
    order: str = Query('desc', description='asc/desc'),
    db: Session = Depends(get_db)
):
    """就业信息列表 (v3j-B-b03 · 支持 search + sort_by + order)"""
    q = db.query(EmploymentRecord).outerjoin(Student, EmploymentRecord.student_id == Student.id)
    if status:
        q = q.filter(EmploymentRecord.status == status)
    if student_id:
        q = q.filter(EmploymentRecord.student_id == student_id)
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Student.name.ilike(pattern),
                Student.student_no.ilike(pattern),
                EmploymentRecord.target_industry.ilike(pattern),
                EmploymentRecord.target_position.ilike(pattern),
                EmploymentRecord.internship_company.ilike(pattern),
                EmploymentRecord.status.ilike(pattern),
            )
        )
    SORT_WHITELIST = {
        'id': EmploymentRecord.id,
        'status': EmploymentRecord.status,
        'sign_date': EmploymentRecord.offer_date,
        'offer_date': EmploymentRecord.offer_date,
        'company': EmploymentRecord.internship_company,
        'position': EmploymentRecord.target_position,
        'student_name': Student.name,
        'student_no': Student.student_no,
    }
    col = SORT_WHITELIST.get(sort_by, EmploymentRecord.id)
    if (order or 'desc').lower() == 'asc':
        q = q.order_by(col.asc())
    else:
        q = q.order_by(col.desc())
    items = q.all()
    return [_emp_dict(e, db) for e in items]


def _emp_export_workbook(items, db):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '就业信息'
    ws.append(['学号', '姓名', '班级', '就业状态', '单位/院校', '岗位/专业', '签约日期', '薪资', '意向类型', '目标行业', '备注'])
    for e in items:
        d = _emp_dict(e, db)
        ws.append([
            d['student_no'], d['student_name'], d['class_name'],
            d['status'], d['company'], d['position'],
            d['sign_date'], d['salary'],
            d.get('intention_type', ''), d.get('target_industry', ''),
            d.get('notes', ''),
        ])
    return wb


@router.post('/employment/export')
def export_employment_by_ids(payload: dict = Body(...), db: Session = Depends(get_db)):
    """按 ID 列表批量导出就业信息 Excel (v3j-B-b03)"""
    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请传入非空的 ids 列表')
    items = db.query(EmploymentRecord).filter(EmploymentRecord.id.in_(ids)).all()
    wb = _emp_export_workbook(items, db)
    return _xlsx_stream(wb, 'employment_selected.xlsx')


@router.get('/employment/export/all')
def export_employment_all(
    search: str = Query(''),
    status: Optional[str] = Query(None),
    student_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """按当前搜索条件导出全部就业信息 Excel (v3j-B-b03)"""
    q = db.query(EmploymentRecord).outerjoin(Student, EmploymentRecord.student_id == Student.id)
    if status:
        q = q.filter(EmploymentRecord.status == status)
    if student_id:
        q = q.filter(EmploymentRecord.student_id == student_id)
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Student.name.ilike(pattern),
                Student.student_no.ilike(pattern),
                EmploymentRecord.target_industry.ilike(pattern),
                EmploymentRecord.target_position.ilike(pattern),
                EmploymentRecord.internship_company.ilike(pattern),
                EmploymentRecord.status.ilike(pattern),
            )
        )
    items = q.order_by(EmploymentRecord.id.desc()).all()
    wb = _emp_export_workbook(items, db)
    return _xlsx_stream(wb, 'employment_all.xlsx')


@router.post('/employment')
def create_employment(data: dict, db: Session = Depends(get_db)):
    e = EmploymentRecord(**data)
    db.add(e)
    db.commit()
    db.refresh(e)
    return {'id': e.id}


@router.put('/employment/{eid}')
def update_employment(eid: int, data: dict, db: Session = Depends(get_db)):
    e = db.query(EmploymentRecord).filter(EmploymentRecord.id == eid).first()
    if not e:
        raise HTTPException(404, '记录不存在')
    for k, v in data.items():
        if hasattr(e, k):
            setattr(e, k, v)
    db.commit()
    return {'ok': True}


@router.delete('/employment/{eid}')
def delete_employment(eid: int, db: Session = Depends(get_db)):
    e = db.query(EmploymentRecord).filter(EmploymentRecord.id == eid).first()
    if e:
        db.delete(e)
        db.commit()
    return {'ok': True}
