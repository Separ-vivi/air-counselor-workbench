"""学期报表导出 Excel 端点"""
import io
import logging
from datetime import datetime
from fastapi import HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, distinct
from sqlalchemy.orm import Session
from models import (
    Student, ClassModel, Major, GradeRecord, WarningRecord,
    PartyProgress, EmploymentRecord, Activity, ActivitySignup,
    PsychologyRecord, StudentDiscipline, StudentHardship,
    StudentGrant, StudentScholarship, StudentLoan, StudentWorkStudy,
    StudentHonor, StudentDormVisit, StudentLeave, StudentDormChat,
    StudentAttendanceException
)
from routers.semester_report_utils import _semester_date_range, _semester_to_academic_year

logger = logging.getLogger(__name__)


def export_semester_report(semester, db: Session):
    """导出学期报表为 Excel（多 Sheet）- 空数据时返回空白模板不报 500"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(500, 'openpyxl 未安装')

    wb = Workbook()

    # 样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    title_font = Font(bold=True, size=14)

    def write_title(ws, title, col=3):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col)
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')

    def write_headers(ws, row, headers):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def safe_count(query_func):
        try:
            return query_func()
        except Exception:
            return 0

    def safe_query(query_func, default=None):
        try:
            return query_func()
        except Exception:
            return default if default is not None else []

    # ==================== Sheet 1: 总览 ====================
    ws1 = wb.active
    ws1.title = '总览'
    write_title(ws1, '学期总览')

    try:
        total_students = db.query(Student).count()
        total_classes = db.query(ClassModel).count()
        total_majors = db.query(Major).count()
    except Exception:
        total_students = total_classes = total_majors = 0

    r = 3
    ws1.cell(row=r, column=1, value='指标')
    ws1.cell(row=r, column=2, value='数值')
    ws1.cell(row=r, column=1).font = header_font
    ws1.cell(row=r, column=2).font = header_font
    for label, val in [('学生总数', total_students), ('班级数', total_classes), ('专业数', total_majors)]:
        r += 1
        ws1.cell(row=r, column=1, value=label)
        ws1.cell(row=r, column=2, value=val)

    # 党员人数（按学期过滤）
    try:
        _exp_pp_subq = db.query(
            PartyProgress.student_id,
            func.max(PartyProgress.id).label('max_id')
        )
        if semester and semester != 'all':
            _exp_start, _exp_end = _semester_date_range(semester)
            if _exp_start and _exp_end:
                _exp_pp_subq = _exp_pp_subq.filter(
                    PartyProgress.stage_date >= _exp_start,
                    PartyProgress.stage_date <= _exp_end
                )
        _exp_pp_subq = _exp_pp_subq.group_by(PartyProgress.student_id).subquery()
        party_member_count = (
            db.query(func.count(PartyProgress.student_id))
            .join(_exp_pp_subq, PartyProgress.id == _exp_pp_subq.c.max_id)
            .filter(PartyProgress.stage.in_(['中共预备党员', '中共党员']))
            .scalar() or 0
        )
        r += 1
        ws1.cell(row=r, column=1, value='党员人数')
        ws1.cell(row=r, column=2, value=party_member_count)
    except Exception:
        r += 1
        ws1.cell(row=r, column=1, value='党员人数')
        ws1.cell(row=r, column=2, value=0)

    # 政治面貌
    r += 2
    ws1.cell(row=r, column=1, value='政治面貌分布')
    ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws1, r, ['政治面貌', '人数'])
    try:
        rows = db.query(Student.political_status, func.count(Student.id)).group_by(Student.political_status).all()
        for status, cnt in rows:
            r += 1
            ws1.cell(row=r, column=1, value=status or '群众')
            ws1.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # 性别
    r += 2
    ws1.cell(row=r, column=1, value='性别分布')
    ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws1, r, ['性别', '人数'])
    try:
        rows = db.query(Student.gender, func.count(Student.id)).group_by(Student.gender).all()
        for g, cnt in rows:
            r += 1
            ws1.cell(row=r, column=1, value=g or '未知')
            ws1.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # 住宿
    r += 2
    ws1.cell(row=r, column=1, value='住宿分布')
    ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws1, r, ['类型', '人数'])
    try:
        off = db.query(Student).filter(Student.is_off_campus == True).count()
        ws1.cell(row=r+1, column=1, value='住校')
        ws1.cell(row=r+1, column=2, value=total_students - off)
        ws1.cell(row=r+2, column=1, value='外宿')
        ws1.cell(row=r+2, column=2, value=off)
        r += 2
    except Exception:
        pass

    # 校区
    r += 2
    ws1.cell(row=r, column=1, value='校区分布')
    ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws1, r, ['校区', '人数'])
    try:
        rows = db.query(Student.campus, func.count(Student.id)).group_by(Student.campus).all()
        for c, cnt in rows:
            r += 1
            ws1.cell(row=r, column=1, value=c or '未知')
            ws1.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # ==================== Sheet 2: 学业 ====================
    ws2 = wb.create_sheet('学业')
    write_title(ws2, '学业数据汇总')

    r = 3
    write_headers(ws2, r, ['班级', '平均成绩'])
    try:
        grade_filters = [GradeRecord.score.isnot(None)]
        if semester and semester != 'all':
            grade_filters.append(GradeRecord.semester == semester)
        rows = (
            db.query(ClassModel.class_name, func.avg(GradeRecord.score))
            .join(Student, Student.class_id == ClassModel.id)
            .join(GradeRecord, GradeRecord.student_id == Student.id)
            .filter(*grade_filters)
            .group_by(ClassModel.id, ClassModel.class_name)
            .all()
        )
        for name, avg in rows:
            r += 1
            ws2.cell(row=r, column=1, value=name)
            ws2.cell(row=r, column=2, value=round(avg, 2) if avg else 0)
    except Exception:
        pass

    # 挂科率
    r += 2
    ws2.cell(row=r, column=1, value='挂科统计')
    ws2.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    try:
        total_with_grades = db.query(func.count(GradeRecord.student_id.distinct())).scalar() or 0
        fail_students = db.query(func.count(GradeRecord.student_id.distinct())).filter(GradeRecord.score.isnot(None), GradeRecord.score < 60).scalar() or 0
        fail_rate = round(fail_students / total_with_grades * 100, 2) if total_with_grades > 0 else 0.0
        ws2.cell(row=r, column=1, value='有成绩学生数')
        ws2.cell(row=r, column=2, value=total_with_grades)
        r += 1
        ws2.cell(row=r, column=1, value='挂科学生数')
        ws2.cell(row=r, column=2, value=fail_students)
        r += 1
        ws2.cell(row=r, column=1, value='挂科率(%)')
        ws2.cell(row=r, column=2, value=fail_rate)
    except Exception:
        pass

    # Top10
    r += 2
    ws2.cell(row=r, column=1, value='成绩排名 Top 10')
    ws2.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws2, r, ['排名', '学号', '姓名', '平均成绩'])
    try:
        rows = (
            db.query(Student.student_no, Student.name, func.avg(GradeRecord.score))
            .join(GradeRecord, GradeRecord.student_id == Student.id)
            .filter(GradeRecord.score.isnot(None))
            .group_by(Student.id, Student.student_no, Student.name)
            .order_by(func.avg(GradeRecord.score).desc())
            .limit(10)
            .all()
        )
        for idx, (sno, sname, avg) in enumerate(rows, 1):
            r += 1
            ws2.cell(row=r, column=1, value=idx)
            ws2.cell(row=r, column=2, value=sno)
            ws2.cell(row=r, column=3, value=sname)
            ws2.cell(row=r, column=4, value=round(avg, 2) if avg else 0)
    except Exception:
        pass

    # 预警
    r += 2
    ws2.cell(row=r, column=1, value='学业预警统计')
    ws2.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws2, r, ['预警等级', '人数'])
    try:
        red = db.query(WarningRecord).filter(WarningRecord.warning_type == 'red').count()
        yellow = db.query(WarningRecord).filter(WarningRecord.warning_type == 'yellow').count()
        r += 1
        ws2.cell(row=r, column=1, value='红色预警')
        ws2.cell(row=r, column=2, value=red)
        r += 1
        ws2.cell(row=r, column=1, value='黄色预警')
        ws2.cell(row=r, column=2, value=yellow)
        r += 1
        ws2.cell(row=r, column=1, value='合计')
        ws2.cell(row=r, column=2, value=red + yellow)
    except Exception:
        pass

    # ==================== Sheet 3: 党团 ====================
    ws3 = wb.create_sheet('党团')
    write_title(ws3, '党团发展进度')
    r = 3
    write_headers(ws3, r, ['阶段', '人数'])
    try:
        subq = (
            db.query(PartyProgress.student_id, func.max(PartyProgress.id).label('max_id'))
            .group_by(PartyProgress.student_id)
            .subquery()
        )
        rows = (
            db.query(PartyProgress.stage, func.count(PartyProgress.student_id))
            .join(subq, PartyProgress.id == subq.c.max_id)
            .group_by(PartyProgress.stage)
            .all()
        )
        for stage, cnt in rows:
            r += 1
            ws3.cell(row=r, column=1, value=stage)
            ws3.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    r += 2
    try:
        now = datetime.now()
        sem_start = f"{now.year}-09-01" if now.month >= 9 else f"{now.year}-03-01"
        new_cnt = db.query(func.count(PartyProgress.id)).filter(PartyProgress.stage_date >= sem_start).scalar() or 0
        ws3.cell(row=r, column=1, value='本学期新发展人数')
        ws3.cell(row=r, column=2, value=new_cnt)
    except Exception:
        pass

    # ==================== Sheet 4: 就业 ====================
    ws4 = wb.create_sheet('就业')
    write_title(ws4, '就业跟踪统计')
    r = 3
    write_headers(ws4, r, ['就业状态', '人数'])
    try:
        rows = db.query(EmploymentRecord.status, func.count(EmploymentRecord.id)).group_by(EmploymentRecord.status).all()
        total_emp = sum(cnt for _, cnt in rows)
        for status, cnt in rows:
            r += 1
            ws4.cell(row=r, column=1, value=status or '未知')
            ws4.cell(row=r, column=2, value=cnt)
        r += 1
        ws4.cell(row=r, column=1, value='合计')
        ws4.cell(row=r, column=2, value=total_emp)
        r += 1
        employed = sum(cnt for s, cnt in rows if s == '已签约')
        rate = round(employed / total_emp * 100, 2) if total_emp > 0 else 0.0
        ws4.cell(row=r, column=1, value='就业率(%)')
        ws4.cell(row=r, column=2, value=rate)
    except Exception:
        pass

    # ==================== Sheet 5: 活动 ====================
    ws5 = wb.create_sheet('活动')
    write_title(ws5, '学生活动统计')
    r = 3
    try:
        total_act = db.query(Activity).count()
        total_ppl = db.query(ActivitySignup).count()
    except Exception:
        total_act = total_ppl = 0
    ws5.cell(row=r, column=1, value='活动总数')
    ws5.cell(row=r, column=2, value=total_act)
    r += 1
    ws5.cell(row=r, column=1, value='参与人次')
    ws5.cell(row=r, column=2, value=total_ppl)
    r += 2
    ws5.cell(row=r, column=1, value='活动参与人数排名')
    ws5.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws5, r, ['活动名称', '类型', '参与人数'])
    try:
        rows = (
            db.query(Activity.title, Activity.activity_type, func.count(ActivitySignup.id))
            .join(ActivitySignup, ActivitySignup.activity_id == Activity.id)
            .group_by(Activity.id, Activity.title, Activity.activity_type)
            .order_by(func.count(ActivitySignup.id).desc())
            .limit(10)
            .all()
        )
        for title, atype, cnt in rows:
            r += 1
            ws5.cell(row=r, column=1, value=title)
            ws5.cell(row=r, column=2, value=atype or '')
            ws5.cell(row=r, column=3, value=cnt)
    except Exception:
        pass

    # ==================== Sheet 6: 考勤 ====================
    ws6 = wb.create_sheet('考勤')
    write_title(ws6, '考勤异常汇总')
    r = 3
    write_headers(ws6, r, ['异常类型', '次数'])
    try:
        q = db.query(
            StudentAttendanceException.exception_type,
            func.count(StudentAttendanceException.id)
        ).group_by(StudentAttendanceException.exception_type)
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    StudentAttendanceException.exception_date >= start,
                    StudentAttendanceException.exception_date <= end
                )
        rows = q.all()
        total_att = 0
        for t, cnt in rows:
            r += 1
            ws6.cell(row=r, column=1, value=t or '未知')
            ws6.cell(row=r, column=2, value=cnt)
            total_att += cnt
        r += 1
        ws6.cell(row=r, column=1, value='合计')
        ws6.cell(row=r, column=2, value=total_att)
    except Exception:
        pass

    # 按班级
    r += 2
    ws6.cell(row=r, column=1, value='按班级统计')
    ws6.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws6, r, ['班级', '异常次数'])
    try:
        q = (
            db.query(ClassModel.class_name, func.count(StudentAttendanceException.id))
            .join(Student, StudentAttendanceException.student_id == Student.id)
            .join(ClassModel, Student.class_id == ClassModel.id)
            .group_by(ClassModel.id, ClassModel.class_name)
        )
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    StudentAttendanceException.exception_date >= start,
                    StudentAttendanceException.exception_date <= end
                )
        rows = q.all()
        for name, cnt in rows:
            r += 1
            ws6.cell(row=r, column=1, value=name)
            ws6.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # ==================== Sheet 7: 心理 ====================
    ws7 = wb.create_sheet('心理')
    write_title(ws7, '心理档案汇总')
    r = 3
    write_headers(ws7, r, ['关注等级', '人数'])
    try:
        rows = (
            db.query(PsychologyRecord.attention_level, func.count(PsychologyRecord.student_id.distinct()))
            .group_by(PsychologyRecord.attention_level)
            .all()
        )
        for level, cnt in rows:
            r += 1
            ws7.cell(row=r, column=1, value=level or '普通')
            ws7.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    r += 2
    try:
        total_counseling = db.query(func.sum(PsychologyRecord.counseling_count)).scalar() or 0
        ws7.cell(row=r, column=1, value='总咨询次数')
        ws7.cell(row=r, column=2, value=total_counseling)
    except Exception:
        pass

    # ==================== Sheet 8: 资助 ====================
    ws8 = wb.create_sheet('资助')
    write_title(ws8, '资助汇总')
    aid_year = _semester_to_academic_year(semester)

    r = 3
    ws8.cell(row=r, column=1, value='困难认定')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws8, r, ['等级', '人数'])
    try:
        q = db.query(StudentHardship.hardship_level, func.count(StudentHardship.id)).group_by(StudentHardship.hardship_level)
        if aid_year:
            q = q.filter(StudentHardship.academic_year == aid_year)
        rows = q.all()
        for level, cnt in rows:
            r += 1
            ws8.cell(row=r, column=1, value=level or '未知')
            ws8.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    r += 2
    ws8.cell(row=r, column=1, value='助学金')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws8, r, ['项目', '数值'])
    try:
        q = db.query(func.sum(StudentGrant.amount), func.count(StudentGrant.id))
        if aid_year:
            q = q.filter(StudentGrant.academic_year == aid_year)
        total_amount, count = q.first()
        r += 1
        ws8.cell(row=r, column=1, value='发放总额')
        ws8.cell(row=r, column=2, value=total_amount or 0)
        r += 1
        ws8.cell(row=r, column=1, value='发放人数')
        ws8.cell(row=r, column=2, value=count or 0)
    except Exception:
        pass

    r += 2
    ws8.cell(row=r, column=1, value='奖学金')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws8, r, ['项目', '数值'])
    try:
        q = db.query(func.sum(StudentScholarship.amount), func.count(StudentScholarship.id))
        if aid_year:
            q = q.filter(StudentScholarship.academic_year == aid_year)
        total_amount, count = q.first()
        r += 1
        ws8.cell(row=r, column=1, value='发放总额')
        ws8.cell(row=r, column=2, value=total_amount or 0)
        r += 1
        ws8.cell(row=r, column=1, value='获奖人数')
        ws8.cell(row=r, column=2, value=count or 0)
    except Exception:
        pass

    r += 2
    ws8.cell(row=r, column=1, value='助学贷款')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    try:
        total_loan = db.query(func.sum(StudentLoan.amount)).scalar() or 0
        loan_cnt = db.query(StudentLoan).count()
        ws8.cell(row=r, column=1, value='贷款总额')
        ws8.cell(row=r, column=2, value=total_loan)
        r += 1
        ws8.cell(row=r, column=1, value='贷款人数')
        ws8.cell(row=r, column=2, value=loan_cnt)
    except Exception:
        pass

    r += 2
    ws8.cell(row=r, column=1, value='勤工助学')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    try:
        q = db.query(func.sum(StudentWorkStudy.compensation), func.count(StudentWorkStudy.id))
        if aid_year:
            q = q.filter(StudentWorkStudy.academic_year == aid_year)
        total_comp, count = q.first()
        ws8.cell(row=r, column=1, value='总报酬')
        ws8.cell(row=r, column=2, value=total_comp or 0)
        r += 1
        ws8.cell(row=r, column=1, value='参与人次')
        ws8.cell(row=r, column=2, value=count or 0)
    except Exception:
        pass

    # ==================== Sheet 9: 荣誉 ====================
    ws9 = wb.create_sheet('荣誉')
    write_title(ws9, '评优评先统计')
    r = 3
    write_headers(ws9, r, ['级别', '获奖人次'])
    try:
        rows = db.query(StudentHonor.level, func.count(StudentHonor.id)).group_by(StudentHonor.level).all()
        for level, cnt in rows:
            r += 1
            ws9.cell(row=r, column=1, value=level or '未知')
            ws9.cell(row=r, column=2, value=cnt)
    except Exception:
        pass
    r += 2
    try:
        student_count = db.query(StudentHonor.student_id).distinct().count()
        ws9.cell(row=r, column=1, value='获奖学生数')
        ws9.cell(row=r, column=2, value=student_count)
    except Exception:
        pass

    # ==================== Sheet 10: 违纪 ====================
    ws10 = wb.create_sheet('违纪')
    write_title(ws10, '违纪处分统计')
    r = 3
    write_headers(ws10, r, ['处分类型', '次数'])
    try:
        rows = db.query(StudentDiscipline.discipline_type, func.count(StudentDiscipline.id)).group_by(StudentDiscipline.discipline_type).all()
        for t, cnt in rows:
            r += 1
            ws10.cell(row=r, column=1, value=t or '未知')
            ws10.cell(row=r, column=2, value=cnt)
    except Exception:
        pass
    r += 2
    try:
        student_count = db.query(StudentDiscipline.student_id).distinct().count()
        ws10.cell(row=r, column=1, value='涉及学生数')
        ws10.cell(row=r, column=2, value=student_count)
    except Exception:
        pass

    # ==================== Sheet 11: 访谈 ====================
    ws11 = wb.create_sheet('访谈')
    write_title(ws11, '学生访谈统计')
    r = 3
    write_headers(ws11, r, ['访谈类型', '次数'])
    try:
        q = db.query(StudentInterview.interview_type, func.count(StudentInterview.id)).group_by(StudentInterview.interview_type)
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(StudentInterview.interview_date >= start, StudentInterview.interview_date <= end)
        rows = q.all()
        for t, cnt in rows:
            r += 1
            ws11.cell(row=r, column=1, value=t or '未知')
            ws11.cell(row=r, column=2, value=cnt)
    except Exception:
        pass
    r += 2
    try:
        q_all = db.query(StudentInterview)
        q_pending = db.query(StudentInterview).filter(StudentInterview.status == '需跟进')
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q_all = q_all.filter(StudentInterview.interview_date >= start, StudentInterview.interview_date <= end)
                q_pending = q_pending.filter(StudentInterview.interview_date >= start, StudentInterview.interview_date <= end)
        total_interviews = q_all.count()
        pending = q_pending.count()
        ws11.cell(row=r, column=1, value='访谈总次数')
        ws11.cell(row=r, column=2, value=total_interviews)
        r += 1
        ws11.cell(row=r, column=1, value='待跟进数量')
        ws11.cell(row=r, column=2, value=pending)
    except Exception:
        pass
    # 访谈覆盖率
    try:
        q_covered = db.query(StudentInterview.student_id).distinct()
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q_covered = q_covered.filter(StudentInterview.interview_date >= start, StudentInterview.interview_date <= end)
        covered_count = q_covered.count()
        total_stu = db.query(Student).count()
        cov_rate = round(covered_count / total_stu * 100, 1) if total_stu > 0 else 0.0
        r += 2
        ws11.cell(row=r, column=1, value='访谈覆盖率(%)')
        ws11.cell(row=r, column=2, value=cov_rate)
        r += 1
        ws11.cell(row=r, column=1, value='被访谈学生数')
        ws11.cell(row=r, column=2, value=covered_count)
        r += 1
        ws11.cell(row=r, column=1, value='总学生数')
        ws11.cell(row=r, column=2, value=total_stu)
    except Exception:
        pass

    # ==================== Sheet 12: 宿舍 ====================
    ws12 = wb.create_sheet('宿舍')
    write_title(ws12, '宿舍管理汇总')
    r = 3
    write_headers(ws12, r, ['项目', '数值'])
    try:
        visit_cnt = db.query(StudentDormVisit).count()
        chat_cnt = db.query(StudentDormChat).count()
        leave_cnt = db.query(StudentLeave).count()
        r += 1
        ws12.cell(row=r, column=1, value='宿舍走访次数')
        ws12.cell(row=r, column=2, value=visit_cnt)
        r += 1
        ws12.cell(row=r, column=1, value='寝谈记录数')
        ws12.cell(row=r, column=2, value=chat_cnt)
        r += 1
        ws12.cell(row=r, column=1, value='请假记录数')
        ws12.cell(row=r, column=2, value=leave_cnt)
    except Exception:
        pass

    # 请假按类型
    r += 2
    ws12.cell(row=r, column=1, value='请假按类型')
    ws12.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws12, r, ['请假类型', '次数'])
    try:
        rows = db.query(StudentLeave.leave_type, func.count(StudentLeave.id)).group_by(StudentLeave.leave_type).all()
        for t, cnt in rows:
            r += 1
            ws12.cell(row=r, column=1, value=t or '未知')
            ws12.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # 写入 buffer
    try:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as e:
        logger.error(f'semester report export save failed: {e}')
        raise HTTPException(500, f'Excel 生成失败: {type(e).__name__}: {str(e)}')

    from urllib.parse import quote
    filename = f"学期报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
