"""活动日程 + 党团学习 + 班会记录 路由"""
from fastapi import APIRouter, Depends, HTTPException, Query, Body
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional
from database import get_db
from models import Activity, ActivitySignup, PartyStudy, ClassMeeting, Student

router = APIRouter(prefix='/api')


# ===== 活动日程 =====
def _act_dict(a):
    return {
        'id': a.id,
        'title': a.title, 'activity_name': a.title,   # alias
        'activity_date': a.activity_date,
        'end_date': a.end_date, 'location': a.location,
        'description': a.description, 'activity_type': a.activity_type,
        'status': a.status, 'max_participants': a.max_participants,
        'organizer': getattr(a, 'organizer', '') or '',
    }


def _act_normalize_input(data: dict) -> dict:
    d = dict(data)
    if 'activity_name' in d and 'title' not in d:
        d['title'] = d.pop('activity_name')
    allowed = {'title','activity_date','end_date','location','description','activity_type','status','max_participants','organizer'}
    return {k: v for k, v in d.items() if k in allowed}


@router.get('/activities')
def list_activities(
    search: str = Query('', description='搜索活动名/类型/组织者/地点'),
    sort_by: str = Query('activity_date', description='排序字段'),
    order: str = Query('desc', description='asc/desc'),
    db: Session = Depends(get_db)
):
    """活动列表 (v3j-B-b02 · 支持 search + sort_by + order)"""
    query = db.query(Activity)
    if search:
        pattern = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Activity.title.ilike(pattern),
                Activity.activity_type.ilike(pattern),
                Activity.organizer.ilike(pattern),
                Activity.location.ilike(pattern),
            )
        )
    SORT_WHITELIST = {
        'activity_name': Activity.title,
        'title': Activity.title,
        'activity_type': Activity.activity_type,
        'activity_date': Activity.activity_date,
        'organizer': Activity.organizer,
        'location': Activity.location,
        'status': Activity.status,
    }
    col = SORT_WHITELIST.get(sort_by, Activity.activity_date)
    if (order or 'desc').lower() == 'asc':
        query = query.order_by(col.asc())
    else:
        query = query.order_by(col.desc())
    items = query.all()
    return [_act_dict(a) for a in items]


@router.post('/activities/export')
def export_activities_by_ids(
    payload: dict = Body(...),
    db: Session = Depends(get_db)
):
    """按 ID 列表批量导出活动 Excel (v3j-B-b02)"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请传入非空的 ids 列表')

    rows = db.query(Activity).filter(Activity.id.in_(ids)).order_by(Activity.activity_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '活动列表'
    ws.append(['活动名称', '类型', '日期', '结束日期', '地点', '组织者', '状态', '说明'])
    for a in rows:
        ws.append([
            a.title, a.activity_type or '', a.activity_date or '',
            a.end_date or '', a.location or '', getattr(a, 'organizer', '') or '',
            a.status or '', a.description or ''
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=activities_selected.xlsx'}
    )


@router.get('/activities/export')
def export_activities_all(
    search: str = Query(''),
    db: Session = Depends(get_db)
):
    """按当前搜索条件导出全部活动 Excel (v3j-B-b02)"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    query = db.query(Activity)
    if search:
        pattern = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Activity.title.ilike(pattern),
                Activity.activity_type.ilike(pattern),
                Activity.organizer.ilike(pattern),
                Activity.location.ilike(pattern),
            )
        )
    rows = query.order_by(Activity.activity_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '活动列表'
    ws.append(['活动名称', '类型', '日期', '结束日期', '地点', '组织者', '状态', '说明'])
    for a in rows:
        ws.append([
            a.title, a.activity_type or '', a.activity_date or '',
            a.end_date or '', a.location or '', getattr(a, 'organizer', '') or '',
            a.status or '', a.description or ''
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=activities_all.xlsx'}
    )


@router.get('/activities/{activity_id}')
def get_activity(activity_id: int, db: Session = Depends(get_db)):
    a = db.query(Activity).get(activity_id)
    if not a:
        raise HTTPException(404, '活动不存在')
    # 获取报名人员
    signups = db.query(ActivitySignup).filter(ActivitySignup.activity_id == activity_id).all()
    participants = []
    for s in signups:
        stu = db.query(Student).get(s.student_id)
        participants.append({
            'id': s.id, 'student_id': s.student_id,
            'student_name': stu.name if stu else '',
            'signed_up': s.signed_up, 'checked_in': s.checked_in,
        })
    _r = _act_dict(a)
    _r['participants'] = participants
    return _r


@router.post('/activities')
def create_activity(data: dict, db: Session = Depends(get_db)):
    a = Activity(**_act_normalize_input(data))
    db.add(a)
    db.commit()
    db.refresh(a)
    return {'id': a.id}


@router.put('/activities/{aid}')
def update_activity(aid: int, data: dict, db: Session = Depends(get_db)):
    a = db.query(Activity).get(aid)
    if not a:
        raise HTTPException(404)
    for k, v in _act_normalize_input(data).items():
        if hasattr(a, k):
            setattr(a, k, v)
    db.commit()
    return {'ok': True}


@router.delete('/activities/{aid}')
def delete_activity(aid: int, db: Session = Depends(get_db)):
    a = db.query(Activity).get(aid)
    if a:
        db.delete(a)
        db.commit()
    return {'ok': True}


@router.get('/activities/{aid}/signups')
def list_activity_signups(aid: int, db: Session = Depends(get_db)):
    from models import ClassModel
    items = db.query(ActivitySignup).filter(ActivitySignup.activity_id == aid).all()
    result = []
    for s in items:
        stu = db.query(Student).filter(Student.id == s.student_id).first()
        cls_name = ''
        if stu and stu.class_id:
            cls = db.query(ClassModel).filter(ClassModel.id == stu.class_id).first()
            cls_name = cls.class_name if cls else ''
        result.append({
            'id': s.id, 'student_id': s.student_id,
            'student_name': stu.name if stu else '',
            'class_name': cls_name,
            'signed_up': s.signed_up, 'checked_in': s.checked_in,
            'points': s.points,
        })
    return result


@router.post('/activities/{aid}/signups')
def create_signup(aid: int, data: dict, db: Session = Depends(get_db)):
    s = ActivitySignup(activity_id=aid, **data)
    db.add(s)
    db.commit()
    return {'id': s.id}


@router.put('/activities/{aid}/signups/{sid}')
def update_signup(aid: int, sid: int, data: dict, db: Session = Depends(get_db)):
    s = db.query(ActivitySignup).filter(ActivitySignup.id == sid, ActivitySignup.activity_id == aid).first()
    if not s:
        raise HTTPException(404)
    for k, v in data.items():
        if hasattr(s, k):
            setattr(s, k, v)
    db.commit()
    return {'ok': True}


# ===== 党团学习 =====
@router.get('/party-study')
def list_party_study(db: Session = Depends(get_db)):
    items = db.query(PartyStudy).order_by(PartyStudy.study_date.desc()).all()
    result = []
    for p in items:
        participants = []
        for stu in p.students:
            participants.append({'id': stu.id, 'name': stu.name})
        result.append({
            'id': p.id, 'study_type': p.study_type, 'study_date': p.study_date,
            'topic': p.topic, 'content_summary': p.content_summary,
            'report_points': p.report_points, 'participant_count': len(participants),
            'participants': participants,
        })
    return result


@router.get('/party-study/{study_id}')
def get_party_study(study_id: int, db: Session = Depends(get_db)):
    p = db.query(PartyStudy).get(study_id)
    if not p:
        raise HTTPException(404, '记录不存在')
    participants = []
    for stu in p.students:
        participants.append({'id': stu.id, 'name': stu.name})
    return {
        'id': p.id, 'study_type': p.study_type, 'study_date': p.study_date,
        'topic': p.topic, 'content_summary': p.content_summary,
        'report_points': p.report_points, 'participant_count': len(participants),
        'participants': participants,
    }


@router.post('/party-study')
def create_party_study(data: dict, db: Session = Depends(get_db)):
    student_ids = data.pop('student_ids', [])
    p = PartyStudy(**data)
    if student_ids:
        students = db.query(Student).filter(Student.id.in_(student_ids)).all()
        p.students = students
    db.add(p)
    db.commit()
    db.refresh(p)
    return {'id': p.id}


@router.put('/party-study/{pid}')
def update_party_study(pid: int, data: dict, db: Session = Depends(get_db)):
    p = db.query(PartyStudy).get(pid)
    if not p:
        raise HTTPException(404)
    student_ids = data.pop('student_ids', None)
    for k, v in data.items():
        if hasattr(p, k):
            setattr(p, k, v)
    if student_ids is not None:
        students = db.query(Student).filter(Student.id.in_(student_ids)).all()
        p.students = students
    db.commit()
    return {'ok': True}


@router.delete('/party-study/{pid}')
def delete_party_study(pid: int, db: Session = Depends(get_db)):
    p = db.query(PartyStudy).get(pid)
    if p:
        db.delete(p)
        db.commit()
    return {'ok': True}


# ===== 班会记录 =====
def _meeting_dict(m, cls_name):
    return {
        'id': m.id, 'class_id': m.class_id, 'class_name': cls_name,
        'meeting_date': m.meeting_date,
        'topic': m.topic, 'theme': m.topic,     # alias for frontend
        'attendance_count': m.attendance_count,
        'absent_students': m.absent_students,
        'content_summary': m.content_summary,
        'summary': m.content_summary,            # alias
        'resolution': m.resolution,
        'host': getattr(m, 'host', '') or '',
        'recorder': getattr(m, 'recorder', '') or '',
        'notes': getattr(m, 'notes', '') or '',
    }


def _meeting_normalize_input(data: dict) -> dict:
    d = dict(data)
    if 'theme' in d and 'topic' not in d:
        d['topic'] = d.pop('theme')
    if 'summary' in d and 'content_summary' not in d:
        d['content_summary'] = d.pop('summary')
    allowed = {'class_id','meeting_date','topic','attendance_count','absent_students','content_summary','resolution','photo','host','recorder','notes'}
    return {k: v for k, v in d.items() if k in allowed}


@router.get('/class-meetings')
def list_class_meetings(
    class_name: Optional[str] = Query(None),
    class_id: Optional[int] = Query(None),
    theme: Optional[str] = Query(None),
    search: str = Query('', description='搜索主题/主持人/记录人/备注'),
    sort_by: str = Query('meeting_date', description='排序字段'),
    order: str = Query('desc', description='asc/desc'),
    db: Session = Depends(get_db)
):
    """班会列表 (v3j-B-b03 · 支持 search + sort_by + order)"""
    from models import ClassModel
    q = db.query(ClassMeeting)
    if class_id:
        q = q.filter(ClassMeeting.class_id == class_id)
    if theme:
        q = q.filter(ClassMeeting.topic.contains(theme))
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                ClassMeeting.topic.ilike(pattern),
                ClassMeeting.host.ilike(pattern),
                ClassMeeting.recorder.ilike(pattern),
                ClassMeeting.content_summary.ilike(pattern),
                ClassMeeting.notes.ilike(pattern),
            )
        )
    SORT_WHITELIST = {
        'meeting_date': ClassMeeting.meeting_date,
        'theme': ClassMeeting.topic,
        'topic': ClassMeeting.topic,
        'host': ClassMeeting.host,
        'recorder': ClassMeeting.recorder,
        'attendance_count': ClassMeeting.attendance_count,
    }
    col = SORT_WHITELIST.get(sort_by, ClassMeeting.meeting_date)
    if (order or 'desc').lower() == 'asc':
        q = q.order_by(col.asc())
    else:
        q = q.order_by(col.desc())
    items = q.all()
    result = []
    for m in items:
        cls_name = ''
        if m.class_id:
            cls = db.query(ClassModel).filter(ClassModel.id == m.class_id).first()
            cls_name = cls.class_name if cls else ''
        if class_name and cls_name != class_name:
            continue
        result.append(_meeting_dict(m, cls_name))
    return result


def _meeting_export_workbook(items, db):
    from openpyxl import Workbook
    from models import ClassModel
    class_map = {c.id: c.class_name for c in db.query(ClassModel).all()}
    wb = Workbook()
    ws = wb.active
    ws.title = '班会记录'
    ws.append(['班会主题', '所属班级', '召开日期', '主持人', '记录人', '出席人数', '内容摘要', '备注'])
    for m in items:
        cls_name = class_map.get(m.class_id, '') if m.class_id else ''
        ws.append([
            m.topic or '', cls_name,
            m.meeting_date or '',
            getattr(m, 'host', '') or '',
            getattr(m, 'recorder', '') or '',
            m.attendance_count or 0,
            m.content_summary or '',
            getattr(m, 'notes', '') or '',
        ])
    return wb


@router.post('/class-meetings/export')
def export_class_meetings_by_ids(payload: dict = Body(...), db: Session = Depends(get_db)):
    """按 ID 列表批量导出班会 Excel (v3j-B-b03)"""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请传入非空的 ids 列表')
    items = db.query(ClassMeeting).filter(ClassMeeting.id.in_(ids)).all()
    wb = _meeting_export_workbook(items, db)
    out = BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(
        out,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=class_meetings_selected.xlsx'}
    )


@router.get('/class-meetings/export/all')
def export_class_meetings_all(
    search: str = Query(''),
    class_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    """按当前搜索条件导出全部班会 Excel (v3j-B-b03)"""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    q = db.query(ClassMeeting)
    if class_id:
        q = q.filter(ClassMeeting.class_id == class_id)
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                ClassMeeting.topic.ilike(pattern),
                ClassMeeting.host.ilike(pattern),
                ClassMeeting.recorder.ilike(pattern),
                ClassMeeting.content_summary.ilike(pattern),
                ClassMeeting.notes.ilike(pattern),
            )
        )
    items = q.order_by(ClassMeeting.meeting_date.desc()).all()
    wb = _meeting_export_workbook(items, db)
    out = BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(
        out,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=class_meetings_all.xlsx'}
    )


@router.get('/class-meetings/{meeting_id}')
def get_class_meeting(meeting_id: int, db: Session = Depends(get_db)):
    from models import ClassModel
    m = db.query(ClassMeeting).filter(ClassMeeting.id == meeting_id).first()
    if not m:
        raise HTTPException(404, '记录不存在')
    cls_name = ''
    if m.class_id:
        cls = db.query(ClassModel).filter(ClassModel.id == m.class_id).first()
        cls_name = cls.class_name if cls else ''
    return _meeting_dict(m, cls_name)


@router.post('/class-meetings')
def create_class_meeting(data: dict, db: Session = Depends(get_db)):
    m = ClassMeeting(**_meeting_normalize_input(data))
    db.add(m)
    db.commit()
    db.refresh(m)
    return {'id': m.id}


@router.put('/class-meetings/{mid}')
def update_class_meeting(mid: int, data: dict, db: Session = Depends(get_db)):
    m = db.query(ClassMeeting).get(mid)
    if not m:
        raise HTTPException(404)
    for k, v in _meeting_normalize_input(data).items():
        if hasattr(m, k):
            setattr(m, k, v)
    db.commit()
    return {'ok': True}


@router.delete('/class-meetings/{mid}')
def delete_class_meeting(mid: int, db: Session = Depends(get_db)):
    m = db.query(ClassMeeting).get(mid)
    if m:
        db.delete(m)
        db.commit()
    return {'ok': True}
