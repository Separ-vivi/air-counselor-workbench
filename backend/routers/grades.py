"""学业预警路由"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
import pandas as pd
import io
from datetime import datetime
from database import get_db
from models import Student, GradeRecord, WarningRecord, Setting, ClassModel
from schemas import GradeOut

router = APIRouter(prefix='/api/grades', tags=['学业管理'])


def _get_settings(db: Session) -> dict:
    """获取预警阈值设置"""
    settings = {s.key: s.value for s in db.query(Setting).all()}
    return {
        'fail_course_threshold': int(settings.get('fail_course_threshold', '2')),
        'gpa_drop_threshold': float(settings.get('gpa_drop_threshold', '0.5')),
    }


def _calculate_warnings(db: Session, student_id: int, semester: str, settings: dict) -> list:
    """计算单个学生在某学期的预警"""
    warnings = []
    grades = db.query(GradeRecord).filter(
        GradeRecord.student_id == student_id, GradeRecord.semester == semester
    ).all()

    if not grades:
        return warnings

    student = db.query(Student).get(student_id)
    if not student:
        return warnings

    # 红灯: 挂科 >= N 门
    fail_count = sum(1 for g in grades if g.score is not None and g.score < 60)
    if fail_count >= settings['fail_course_threshold']:
        fail_courses = [g.course_name for g in grades if g.score is not None and g.score < 60]
        warnings.append({
            'student_id': student_id,
            'warning_type': 'red',
            'description': f'挂科{fail_count}门: {", ".join(fail_courses)}',
            'semester': semester,
        })

    # 黄灯: 绩点较上学期下降 >= N
    prev_semesters = db.query(GradeRecord.semester).filter(
        GradeRecord.student_id == student_id, GradeRecord.semester < semester
    ).distinct().order_by(GradeRecord.semester.desc()).all()

    if prev_semesters:
        prev_sem = prev_semesters[0][0]
        prev_grades = db.query(GradeRecord).filter(
            GradeRecord.student_id == student_id, GradeRecord.semester == prev_sem
        ).all()

        def calc_gpa(grade_list):
            total_credit = sum(g.credit for g in grade_list if g.credit and g.gpa is not None)
            if total_credit == 0:
                return None
            return sum(g.gpa * g.credit for g in grade_list if g.credit and g.gpa is not None) / total_credit

        curr_gpa = calc_gpa(grades)
        prev_gpa = calc_gpa(prev_grades)

        if curr_gpa is not None and prev_gpa is not None:
            drop = prev_gpa - curr_gpa
            if drop >= settings['gpa_drop_threshold']:
                warnings.append({
                    'student_id': student_id,
                    'warning_type': 'yellow',
                    'description': f'绩点从{prev_gpa:.2f}下降至{curr_gpa:.2f} (下降{drop:.2f})',
                    'semester': semester,
                })

    return warnings


@router.post('/import', response_model=dict)
async def import_grades(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """批量导入成绩 (Excel/CSV)"""
    content = await file.read()
    filename = file.filename or ''

    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f'文件解析失败: {str(e)}')

    # 列名映射
    column_map = {
        '学号': 'student_no', 'student_no': 'student_no',
        '学期': 'semester', 'semester': 'semester',
        '课程': 'course_name', '课程名': 'course_name', '课程名称': 'course_name',
        'course_name': 'course_name',
        '分数': 'score', '成绩': 'score', 'score': 'score', '考试分数': 'score',
        '绩点': 'gpa', 'gpa': 'gpa', 'GPA': 'gpa',
        '学分': 'credit', 'credit': 'credit',
    }

    rename_map = {}
    for col in df.columns:
        col_stripped = str(col).strip()
        if col_stripped in column_map:
            rename_map[col] = column_map[col_stripped]
    df = df.rename(columns=rename_map)

    required = ['student_no', 'semester', 'course_name']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise HTTPException(400, f'文件缺少必要列: {", ".join(missing)}')

    df = df.fillna(0)
    settings = _get_settings(db)
    success_count = 0
    error_count = 0
    warning_count = 0

    for idx, row in df.iterrows():
        try:
            student_no = str(row.get('student_no', '')).strip()
            student = db.query(Student).filter(Student.student_no == student_no).first()
            if not student:
                error_count += 1
                continue

            score_val = float(row.get('score', 0)) if row.get('score', 0) != '' else None
            gpa_val = float(row.get('gpa', 0)) if row.get('gpa', 0) != '' else None
            credit_val = float(row.get('credit', 0)) if row.get('credit', 0) != '' else None

            grade = GradeRecord(
                student_id=student.id,
                semester=str(row.get('semester', '')).strip(),
                course_name=str(row.get('course_name', '')).strip(),
                score=score_val,
                gpa=gpa_val,
                credit=credit_val,
            )
            db.add(grade)
            success_count += 1
        except Exception as e:
            error_count += 1

    db.flush()

    # 导入完成后计算预警
    semesters = df['semester'].unique().tolist()
    student_ids = set()
    for idx, row in df.iterrows():
        student_no = str(row.get('student_no', '')).strip()
        student = db.query(Student).filter(Student.student_no == student_no).first()
        if student:
            student_ids.add(student.id)

    for sid in student_ids:
        for sem in semesters:
            warnings = _calculate_warnings(db, sid, sem, settings)
            for w in warnings:
                # 检查是否已存在相同预警
                existing = db.query(WarningRecord).filter(
                    WarningRecord.student_id == w['student_id'],
                    WarningRecord.warning_type == w['warning_type'],
                    WarningRecord.semester == w['semester'],
                ).first()
                if not existing:
                    record = WarningRecord(**w)
                    db.add(record)
                    warning_count += 1

    db.commit()
    return {
        'message': f'导入完成: 成功{success_count}条, 失败{error_count}条, 新增预警{warning_count}条',
        'success_count': success_count,
        'error_count': error_count,
        'warning_count': warning_count,
    }


def _calc_grade_level(score):
    if score is None: return ''
    if score >= 90: return 'A'
    if score >= 85: return 'A-'
    if score >= 80: return 'B+'
    if score >= 75: return 'B'
    if score >= 70: return 'C+'
    if score >= 65: return 'C'
    if score >= 60: return 'D'
    return 'F'


def _grade_dict(g):
    return {
        'id': g.id, 'semester': g.semester,
        'course_code': getattr(g, 'course_code', '') or '',
        'course_name': g.course_name,
        'score': g.score, 'gpa': g.gpa, 'credit': g.credit,
        'grade_level': getattr(g, 'grade_level', '') or _calc_grade_level(g.score),
        'is_makeup': bool(getattr(g, 'is_makeup', False) or getattr(g, 'is_repair', False)),
        'is_repair': bool(getattr(g, 'is_repair', False)),
    }


@router.get('/student/{student_id}')
def get_student_grades(student_id: int, db: Session = Depends(get_db)):
    """获取学生成绩列表"""
    grades = db.query(GradeRecord).filter(GradeRecord.student_id == student_id).order_by(
        GradeRecord.semester.desc(), GradeRecord.course_name
    ).all()
    return [_grade_dict(g) for g in grades]


@router.get('/semesters')
def get_semesters(db: Session = Depends(get_db)):
    """获取所有学期列表"""
    semesters = db.query(GradeRecord.semester).distinct().order_by(GradeRecord.semester.desc()).all()
    return [s[0] for s in semesters]


@router.get('/warnings')
def get_warnings(
    warning_type: Optional[str] = Query(None, description='预警类型: red/yellow'),
    semester: Optional[str] = Query(None, description='学期筛选'),
    search: str = Query('', description='搜索学号/姓名/班级/预警原因'),
    sort_by: str = Query('warning_type', description='排序字段'),
    order: str = Query('desc', description='asc/desc'),
    db: Session = Depends(get_db)
):
    """获取预警学生列表 (v3j-B-b03 · 支持 search + sort_by + order)"""
    from sqlalchemy import or_ as _or
    query = db.query(WarningRecord, Student).join(
        Student, WarningRecord.student_id == Student.id
    ).outerjoin(ClassModel, Student.class_id == ClassModel.id)
    if warning_type:
        query = query.filter(WarningRecord.warning_type == warning_type)
    if semester:
        query = query.filter(WarningRecord.semester == semester)
    if search:
        pattern = f"%{search.strip()}%"
        query = query.filter(
            _or(
                Student.name.ilike(pattern),
                Student.student_no.ilike(pattern),
                ClassModel.class_name.ilike(pattern),
                WarningRecord.description.ilike(pattern),
                WarningRecord.warning_type.ilike(pattern),
                WarningRecord.semester.ilike(pattern),
            )
        )
    SORT_WHITELIST = {
        'warning_type': WarningRecord.warning_type,
        'warning_level': WarningRecord.warning_type,
        'semester': WarningRecord.semester,
        'created_at': WarningRecord.created_at,
        'student_name': Student.name,
        'student_no': Student.student_no,
        'class_name': ClassModel.class_name,
    }
    col = SORT_WHITELIST.get(sort_by, WarningRecord.warning_type)
    if (order or 'desc').lower() == 'asc':
        query = query.order_by(col.asc(), WarningRecord.created_at.desc())
    else:
        query = query.order_by(col.desc(), WarningRecord.created_at.desc())
    results = query.all()

    # 预先算好每个学生的挂科数和 GPA（当前学期或全部）
    def stats_for(sid, sem):
        q = db.query(GradeRecord).filter(GradeRecord.student_id == sid)
        if sem:
            q = q.filter(GradeRecord.semester == sem)
        grades = q.all()
        scores = [float(g.score) for g in grades if g.score is not None]
        fc = sum(1 for x in scores if x < 60)
        gpa_vals = [float(g.gpa) for g in grades if g.gpa is not None]
        avg_gpa = round(sum(gpa_vals) / len(gpa_vals), 2) if gpa_vals else 0
        return fc, avg_gpa

    def level_label(wt):
        if not wt: return ''
        wt = str(wt)
        if 'red' in wt.lower() or '红' in wt: return '红色'
        if 'yellow' in wt.lower() or '黄' in wt: return '黄色'
        if 'blue' in wt.lower() or '蓝' in wt: return '蓝色'
        return wt

    payload = []
    for w, s in results:
        fc, gpa_avg = stats_for(w.student_id, w.semester)
        payload.append({
            'id': w.id,
            'student_id': w.student_id,
            'warning_type': w.warning_type,
            'warning_level': level_label(w.warning_type),  # 前端 UI 期望
            'description': w.description,
            'warning_reason': w.description or '',        # 前端 UI 期望
            'fail_count': fc,
            'gpa': gpa_avg,
            'semester': w.semester,
            'student_name': s.name,
            'student_no': s.student_no,
            'class_name': s.class_obj.class_name if s.class_obj else '',
            'created_at': w.created_at.isoformat() if w.created_at else None,
            'reminded': bool(w.reminded) if hasattr(w, 'reminded') else False,
            'reminded_at': w.reminded_at.isoformat() if getattr(w, 'reminded_at', None) else None,
        })
    return payload


@router.patch('/warnings/{warning_id}/toggle-reminded')
def toggle_warning_reminded(warning_id: int, db: Session = Depends(get_db)):
    """v3j-D · D1: 切换预警 已提醒 状态"""
    w = db.query(WarningRecord).filter(WarningRecord.id == warning_id).first()
    if not w:
        raise HTTPException(status_code=404, detail='预警记录不存在')
    w.reminded = not bool(w.reminded)
    w.reminded_at = datetime.now() if w.reminded else None
    db.commit()
    return {
        'id': w.id,
        'reminded': bool(w.reminded),
        'reminded_at': w.reminded_at.isoformat() if w.reminded_at else None,
    }


@router.post('/recalculate', response_model=dict)
def recalculate_warnings(db: Session = Depends(get_db)):
    """重新计算所有预警"""
    settings = _get_settings(db)

    # 清除旧预警
    db.query(WarningRecord).delete()
    db.flush()

    # 获取所有有成绩的学生和学期
    student_semesters = db.query(
        GradeRecord.student_id, GradeRecord.semester
    ).distinct().all()

    warning_count = 0
    for student_id, semester in student_semesters:
        warnings = _calculate_warnings(db, student_id, semester, settings)
        for w in warnings:
            record = WarningRecord(**w)
            db.add(record)
            warning_count += 1

    db.commit()
    return {'message': f'预警重新计算完成, 共{warning_count}条预警', 'warning_count': warning_count}


@router.get('/export')
def export_grades(
    student_id: int = None,
    semester: str = None,
    db: Session = Depends(get_db)
):
    """导出成绩 Excel"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    query = db.query(GradeRecord, Student).join(Student, GradeRecord.student_id == Student.id)
    if student_id:
        query = query.filter(GradeRecord.student_id == student_id)
    if semester:
        query = query.filter(GradeRecord.semester == semester)
    
    results = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = '成绩列表'
    
    # 写入表头
    headers = ['学号', '姓名', '班级', '专业', '学期', '课程', '成绩', '绩点']
    ws.append(headers)
    
    # 写入数据
    for grade, student in results:
        cls_name = student.class_obj.class_name if student.class_obj else ''
        major_name = ''
        if student.class_obj and student.class_obj.major:
            major_name = student.class_obj.major.major_name
        ws.append([
            student.student_no,
            student.name,
            cls_name,
            major_name,
            grade.semester,
            grade.course_name,
            grade.score,
            grade.gpa
        ])
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'grades_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@router.post('/export')
def export_grades_by_ids(
    payload: dict = Body(...),
    db: Session = Depends(get_db)
):
    """按 GradeRecord ID 列表批量导出成绩 Excel (v3j-B-b02)"""
    from openpyxl import Workbook
    from io import BytesIO

    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请传入非空的 ids 列表')

    query = db.query(GradeRecord, Student).join(Student, GradeRecord.student_id == Student.id).filter(GradeRecord.id.in_(ids))
    results = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = '成绩列表'
    ws.append(['学号', '姓名', '班级', '专业', '学期', '课程', '成绩', '绩点'])
    for grade, student in results:
        cls_name = student.class_obj.class_name if student.class_obj else ''
        major_name = ''
        if student.class_obj and student.class_obj.major:
            major_name = student.class_obj.major.major_name
        ws.append([
            student.student_no,
            student.name,
            cls_name,
            major_name,
            grade.semester,
            grade.course_name,
            grade.score,
            grade.gpa,
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'grades_selected_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@router.post('/warnings/export/by-ids')
def export_warnings_by_ids(payload: dict = Body(...), db: Session = Depends(get_db)):
    """按 ID 列表批量导出预警 Excel (v3j-B-b03)"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请传入非空的 ids 列表')
    results = db.query(WarningRecord, Student).join(
        Student, WarningRecord.student_id == Student.id
    ).outerjoin(ClassModel, Student.class_id == ClassModel.id).filter(
        WarningRecord.id.in_(ids)
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '预警_选中'
    ws.append(['学号', '姓名', '班级', '专业', '预警类型', '预警描述', '学期', '创建时间'])
    type_map = {'red': '红灯', 'yellow': '黄灯', 'green': '绿灯'}
    for warning, student in results:
        created_at = warning.created_at
        if hasattr(created_at, 'strftime'):
            created_at_str = created_at.strftime('%Y-%m-%d %H:%M')
        else:
            created_at_str = str(created_at) if created_at else ''
        cls_name = student.class_obj.class_name if student.class_obj else ''
        major_name = ''
        if student.class_obj and student.class_obj.major:
            major_name = student.class_obj.major.major_name
        ws.append([
            student.student_no, student.name, cls_name, major_name,
            type_map.get(warning.warning_type, warning.warning_type),
            warning.description, warning.semester, created_at_str,
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'warnings_selected_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@router.get('/warnings/export')
def export_warnings(
    warning_type: str = None,
    class_name: str = None,
    db: Session = Depends(get_db)
):
    """导出预警 Excel"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    query = db.query(WarningRecord, Student).join(Student, WarningRecord.student_id == Student.id).outerjoin(ClassModel, Student.class_id == ClassModel.id)
    if warning_type:
        query = query.filter(WarningRecord.warning_type == warning_type)
    if class_name:
        query = query.filter(ClassModel.class_name == class_name)
    
    results = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = '学业预警'
    
    # 写入表头
    headers = ['学号', '姓名', '班级', '专业', '预警类型', '预警描述', '学期', '创建时间']
    ws.append(headers)
    
    # 写入数据
    type_map = {'red': '红灯', 'yellow': '黄灯', 'green': '绿灯'}
    for warning, student in results:
        created_at = warning.created_at
        if hasattr(created_at, 'strftime'):
            created_at_str = created_at.strftime('%Y-%m-%d %H:%M')
        else:
            created_at_str = str(created_at) if created_at else ''
        cls_name = student.class_obj.class_name if student.class_obj else ''
        major_name = ''
        if student.class_obj and student.class_obj.major:
            major_name = student.class_obj.major.major_name
        ws.append([
            student.student_no,
            student.name,
            cls_name,
            major_name,
            type_map.get(warning.warning_type, warning.warning_type),
            warning.description,
            warning.semester,
            created_at_str
        ])
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'warnings_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@router.get('/by-class/{class_id}')
def get_grades_by_class(
    class_id: int,
    semester: str = None,
    db: Session = Depends(get_db)
):
    """按班级维度查看成绩：返回本班全体学生的成绩汇总统计 + 明细"""
    from models import ClassModel
    cls = db.query(ClassModel).filter(ClassModel.id == class_id).first()
    if not cls:
        return {'class': None, 'students': [], 'grades': [], 'stats': {}}

    students = db.query(Student).filter(Student.class_id == class_id).all()
    sid_list = [s.id for s in students]
    if not sid_list:
        return {
            'class': {'id': cls.id, 'class_name': cls.class_name},
            'students': [], 'grades': [], 'stats': {'total_students': 0, 'total_courses': 0, 'avg_score': 0}
        }

    q = db.query(GradeRecord).filter(GradeRecord.student_id.in_(sid_list))
    if semester:
        q = q.filter(GradeRecord.semester == semester)
    all_grades = q.all()

    # 学生级别汇总
    student_summary = []
    sid_to_name = {s.id: (s.name, s.student_no) for s in students}
    per_student = {}
    for g in all_grades:
        per_student.setdefault(g.student_id, []).append(g)
    for s in students:
        gs = per_student.get(s.id, [])
        scores = [float(g.score) for g in gs if g.score is not None]
        avg = round(sum(scores) / len(scores), 2) if scores else 0
        fail_cnt = sum(1 for x in scores if x < 60)
        student_summary.append({
            'student_id': s.id,
            'student_no': s.student_no,
            'name': s.name,
            'total_courses': len(gs),
            'avg_score': avg,
            'fail_count': fail_cnt,
            'pass_rate': round((len(scores) - fail_cnt) / len(scores) * 100, 1) if scores else 0,
        })

    # 明细列表（供表格展示）
    detail = []
    for g in all_grades:
        nm, no = sid_to_name.get(g.student_id, ('', ''))
        _d = _grade_dict(g)
        _d.update({'student_id': g.student_id, 'student_no': no, 'student_name': nm})
        detail.append(_d)

    all_scores = [float(g.score) for g in all_grades if g.score is not None]
    stats = {
        'total_students': len(students),
        'total_courses': len(all_grades),
        'avg_score': round(sum(all_scores) / len(all_scores), 2) if all_scores else 0,
        'pass_count': sum(1 for x in all_scores if x >= 60),
        'fail_count': sum(1 for x in all_scores if x < 60),
    }

    return {
        'class': {'id': cls.id, 'class_name': cls.class_name},
        'students': student_summary,
        'grades': detail,
        'stats': stats,
    }

