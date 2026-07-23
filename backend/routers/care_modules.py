"""党团发展 + 心理关怀 + 家校沟通 路由"""
from fastapi import APIRouter, Depends, HTTPException, Query, Body
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import Optional
from datetime import datetime, timedelta
from collections import Counter
import json as _json
from database import get_db
from models import PartyProgress, PsychologyRecord, FamilyContact, Student, ClassModel

router = APIRouter(prefix='/api')


# ===== 党团发展 =====
def _party_row_dict(p, stu, class_map):
    return {
        'id': p.id, 'student_id': p.student_id, 'stage': p.stage,
        'stage_date': p.stage_date, 'contact_person': p.contact_person,
        'notes': p.notes, 'student_name': stu.name if stu else '',
        'student_no': stu.student_no if stu else '',
        'class_name': class_map.get(stu.class_id, '') if stu else '',
    }


@router.get('/party-progress')
def list_party_progress(
    search: str = Query('', description='搜索学号/姓名/阶段/联系人'),
    student_id: Optional[int] = Query(None),
    stage: Optional[str] = Query(None),
    class_id: Optional[int] = Query(None, description='v3j-C c01 · 按班级筛选'),
    semester: Optional[str] = Query(None, description='按学期筛选，格式如 2024-2025-1'),
    sort_by: str = Query('stage_date', description='排序字段'),
    order: str = Query('desc', description='asc/desc'),
    db: Session = Depends(get_db)
):
    """党团发展列表 (v3j-B-b03 · 支持 search + sort_by + order；v3j-C c01 · 加 class_id；支持学期筛选)"""
    q = db.query(PartyProgress, Student).join(Student, PartyProgress.student_id == Student.id)
    if student_id:
        q = q.filter(PartyProgress.student_id == student_id)
    if stage:
        q = q.filter(PartyProgress.stage == stage)
    if class_id is not None:
        q = q.filter(Student.class_id == class_id)
    if semester and semester != 'all':
        # 学期转日期范围
        parts = semester.split('-')
        if len(parts) == 3:
            y1, y2, term = parts[0], parts[1], parts[2]
            if term == '1':
                start, end = f"{y1}-09-01", f"{y2}-01-31"
            else:
                start, end = f"{y1}-09-01", f"{y2}-07-31"
            q = q.filter(
                PartyProgress.stage_date >= start,
                PartyProgress.stage_date <= end
            )
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Student.name.ilike(pattern),
                Student.student_no.ilike(pattern),
                PartyProgress.stage.ilike(pattern),
                PartyProgress.contact_person.ilike(pattern),
                PartyProgress.notes.ilike(pattern),
            )
        )
    SORT_WHITELIST = {
        'stage_date': PartyProgress.stage_date,
        'stage': PartyProgress.stage,
        'contact_person': PartyProgress.contact_person,
        'student_name': Student.name,
        'student_no': Student.student_no,
    }
    col = SORT_WHITELIST.get(sort_by, PartyProgress.stage_date)
    if (order or 'desc').lower() == 'asc':
        q = q.order_by(col.asc())
    else:
        q = q.order_by(col.desc())
    class_map = {c.id: c.class_name for c in db.query(ClassModel).all()}
    result = []
    for p, stu in q.all():
        result.append(_party_row_dict(p, stu, class_map))
    return result


@router.post('/party-progress/export-by-ids')
def export_party_progress_by_ids(payload: dict = Body(...), db: Session = Depends(get_db)):
    """按 ID 列表批量导出党团发展 Excel (v3j-B-b03)"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from datetime import datetime as _dt

    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请传入非空的 ids 列表')

    rows = db.query(PartyProgress, Student).join(
        Student, PartyProgress.student_id == Student.id
    ).filter(PartyProgress.id.in_(ids)).all()
    class_map = {c.id: c.class_name for c in db.query(ClassModel).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = '党团发展_选中'
    ws.append(['学号', '姓名', '班级', '发展阶段', '阶段日期', '联系人', '备注'])
    for p, stu in rows:
        ws.append([
            stu.student_no if stu else '',
            stu.name if stu else '',
            class_map.get(stu.class_id, '') if stu else '',
            p.stage or '', p.stage_date or '',
            p.contact_person or '', p.notes or '',
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'party_progress_selected_{_dt.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@router.post('/party-progress')
def create_party_progress(data: dict, db: Session = Depends(get_db)):
    p = PartyProgress(**data)
    db.add(p)
    db.commit()
    db.refresh(p)
    return {'id': p.id}


@router.put('/party-progress/{pid}')
def update_party_progress(pid: int, data: dict, db: Session = Depends(get_db)):
    p = db.get(PartyProgress, pid)
    if not p:
        raise HTTPException(404, '记录不存在')
    for k, v in data.items():
        if hasattr(p, k):
            setattr(p, k, v)
    db.commit()
    return {'ok': True}


@router.delete('/party-progress/{pid}')
def delete_party_progress(pid: int, db: Session = Depends(get_db)):
    p = db.get(PartyProgress, pid)
    if p:
        db.delete(p)
        db.commit()
    return {'ok': True}


@router.get('/party-progress/overview')
def party_progress_overview(
    class_id: Optional[int] = None,
    class_name: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """党团发展全景表 - 支持按 class_id 或 class_name 筛选（含'群众'状态）"""
    stages = ['群众', '递交入党申请书', '团员', '积极分子', '发展对象', '中共预备党员', '中共党员']
    q = db.query(Student)
    if class_id is not None:
        q = q.filter(Student.class_id == class_id)
    elif class_name:
        # 兼容旧调用：class_name → class_id 转换
        target = db.query(ClassModel).filter(ClassModel.class_name == class_name).first()
        if target:
            q = q.filter(Student.class_id == target.id)
        else:
            return {'stages': stages, 'students': []}
    students = q.all()

    # 预加载班级映射，避免 N+1
    class_map = {c.id: c for c in db.query(ClassModel).all()}

    result = []
    for s in students:
        progress_list = db.query(PartyProgress).filter(
            PartyProgress.student_id == s.id
        ).order_by(PartyProgress.created_at.desc()).all()
        # 默认阶段：政治面貌里含"党员/团员/群众"就用该值，否则未记录
        if progress_list:
            current_stage = progress_list[0].stage
        elif s.political_status:
            current_stage = s.political_status
        else:
            current_stage = '群众'
        cls = class_map.get(s.class_id)
        result.append({
            'student_id': s.id, 'name': s.name, 'student_no': s.student_no,
            'class_id': s.class_id,
            'class_name': cls.class_name if cls else '',
            'current_stage': current_stage,
            'political_status': s.political_status or '群众',
            'progress_count': len(progress_list),
        })
    return {'stages': stages, 'students': result}


@router.get('/party-progress/detail/{student_id}')
def party_progress_detail(student_id: int, db: Session = Depends(get_db)):
    """党团发展详情 - 时间线视图"""
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(404, '学生不存在')
    
    progress_list = db.query(PartyProgress).filter(
        PartyProgress.student_id == student_id
    ).order_by(PartyProgress.stage_date.asc()).all()
    
    # 6大阶段定义
    stage_definitions = [
        {'key': '递交入党申请书', 'label': '递交入党申请书', 'description': '提交书面入党申请'},
        {'key': '积极分子', 'label': '确定为入党积极分子', 'description': '经团组织推优确定'},
        {'key': '发展对象', 'label': '确定为发展对象', 'description': '经过一年以上培养考察'},
        {'key': '中共预备党员', 'label': '接收为中共预备党员', 'description': '支部大会讨论通过'},
        {'key': '转正', 'label': '预备党员转正', 'description': '预备期满一年转正'},
        {'key': '中共党员', 'label': '中共党员', 'description': '完成全部发展流程'},
    ]
    
    # 构建时间线
    timeline = []
    progress_map = {p.stage: p for p in progress_list}
    current_stage_index = -1
    
    for i, stage_def in enumerate(stage_definitions):
        progress = progress_map.get(stage_def['key'])
        stage_info = {
            'stage': stage_def['key'],
            'label': stage_def['label'],
            'description': stage_def['description'],
            'completed': progress is not None,
            'date': progress.stage_date if progress else None,
            'contact_person': progress.contact_person if progress else None,
            'notes': progress.notes if progress else None,
            'is_current': False,
        }
        if progress:
            current_stage_index = i
        timeline.append(stage_info)
    
    # 标记当前阶段
    if current_stage_index >= 0 and current_stage_index < len(timeline):
        timeline[current_stage_index]['is_current'] = True
    
    # 获取班级/专业名（Student 没有 class_name/major 字段，需通过 relationship）
    cls_obj = db.get(ClassModel, student.class_id) if student.class_id else None
    major_obj = cls_obj.major if cls_obj else None
    return {
        'student': {
            'id': student.id,
            'name': student.name,
            'student_no': student.student_no,
            'class_name': cls_obj.class_name if cls_obj else '',
            'major': major_obj.major_name if major_obj else '',
        },
        'current_stage': progress_list[-1].stage if progress_list else '未记录',
        'timeline': timeline,
        'records': [
            {
                'id': p.id,
                'stage': p.stage,
                'stage_date': p.stage_date,
                'contact_person': p.contact_person,
                'notes': p.notes,
            }
            for p in progress_list
        ],
    }


@router.get('/party-progress/export')
def export_party_progress(
    stage: str = None,
    class_id: int = None,
    class_name: str = None,
    student_id: int = None,
    search: str = '',
    db: Session = Depends(get_db)
):
    """导出党团发展 Excel（v3j-B b03-batch02 · 支持 search / student_id / stage / class_id / class_name 筛选）"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    from sqlalchemy import or_
    
    query = db.query(PartyProgress, Student).join(Student, PartyProgress.student_id == Student.id)
    if stage:
        query = query.filter(PartyProgress.stage == stage)
    if student_id is not None:
        query = query.filter(PartyProgress.student_id == student_id)
    if class_id is not None:
        query = query.filter(Student.class_id == class_id)
    elif class_name:
        target = db.query(ClassModel).filter(ClassModel.class_name == class_name).first()
        if target:
            query = query.filter(Student.class_id == target.id)
    if search:
        pattern = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Student.name.ilike(pattern),
                Student.student_no.ilike(pattern),
                PartyProgress.stage.ilike(pattern),
                PartyProgress.contact_person.ilike(pattern),
                PartyProgress.notes.ilike(pattern),
            )
        )
    
    results = query.order_by(PartyProgress.stage_date.desc()).all()
    # 预加载班级名映射
    class_map = {c.id: c.class_name for c in db.query(ClassModel).all()}
    
    wb = Workbook()
    ws = wb.active
    ws.title = '党团发展'
    
    # 写入表头
    headers = ['学号', '姓名', '班级', '当前阶段', '阶段日期', '联系人', '备注']
    ws.append(headers)
    
    # 写入数据
    stage_map = {
        'application': '递交入党申请书',
        'activist': '入党积极分子',
        'development': '发展对象',
        'probation': '中共预备党员',
        'regularization': '预备党员转正',
        'full_member': '中共党员'
    }
    for progress, student in results:
        stage_date_str = progress.stage_date
        if hasattr(progress.stage_date, 'strftime'):
            stage_date_str = progress.stage_date.strftime('%Y-%m-%d')
        ws.append([
            student.student_no,
            student.name,
            class_map.get(student.class_id, ''),
            stage_map.get(progress.stage, progress.stage),
            stage_date_str or '',
            progress.contact_person or '',
            progress.notes or ''
        ])
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'party_progress_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ===== 党团发展 统计图表 =====
@router.get('/party-progress/chart-data')
def party_progress_chart_data(semester: Optional[str] = Query(None, description='按学期筛选'), db: Session = Depends(get_db)):
    """党团发展统计图表数据（支持学期筛选）"""
    # 构建学期日期范围过滤
    stage_date_filter = []
    if semester and semester != 'all':
        parts = semester.split('-')
        if len(parts) == 3:
            y1, y2, term = parts[0], parts[1], parts[2]
            if term == '1':
                start, end = f"{y1}-09-01", f"{y2}-01-31"
            else:
                start, end = f"{y1}-09-01", f"{y2}-07-31"
            stage_date_filter = [
                PartyProgress.stage_date >= start,
                PartyProgress.stage_date <= end
            ]
    
    # 1. stage_distribution: GROUP BY stage, COUNT(*)
    stage_q = db.query(PartyProgress.stage, func.count(PartyProgress.id))
    if stage_date_filter:
        stage_q = stage_q.filter(*stage_date_filter)
    stage_rows = stage_q.group_by(PartyProgress.stage).all()
    stage_distribution = [{'stage': s or '', 'count': c} for s, c in stage_rows]

    # 2. monthly_trend: 按 stage_date 的 YYYY-MM 分组，最近12个月
    today = datetime.now()
    months = [(today - timedelta(days=30 * i)).strftime('%Y-%m') for i in range(11, -1, -1)]
    # stage_date 是 String(20)，格式 YYYY-MM-DD，取前7位
    month_q = db.query(
        func.substr(PartyProgress.stage_date, 1, 7).label('month'),
        func.count(PartyProgress.id)
    ).filter(
        PartyProgress.stage_date != '', PartyProgress.stage_date.isnot(None)
    )
    if stage_date_filter:
        month_q = month_q.filter(*stage_date_filter)
    month_rows = month_q.group_by(
        func.substr(PartyProgress.stage_date, 1, 7)
    ).all()
    month_map = {m: c for m, c in month_rows if m}
    monthly_trend = [{'month': m, 'count': month_map.get(m, 0)} for m in months]

    # 3. top_students: 按 student_id 分组，取 TOP10
    top_q = db.query(
        PartyProgress.student_id, func.count(PartyProgress.id).label('cnt')
    )
    if stage_date_filter:
        top_q = top_q.filter(*stage_date_filter)
    top_rows = top_q.group_by(PartyProgress.student_id).order_by(func.count(PartyProgress.id).desc()).limit(10).all()
    top_students = []
    for sid, cnt in top_rows:
        stu = db.get(Student, sid)
        top_students.append({
            'student_name': stu.name if stu else '',
            'student_no': stu.student_no if stu else '',
            'count': cnt,
        })

    return {
        'stage_distribution': stage_distribution,
        'monthly_trend': monthly_trend,
        'top_students': top_students,
    }


# ===== 心理关怀 =====
def _psy_dict(r, db):
    from models import ClassModel
    stu = db.get(Student, r.student_id)
    cls_name = ''
    if stu and stu.class_id:
        cls = db.get(ClassModel, stu.class_id)
        cls_name = cls.class_name if cls else ''
    return {
        'id': r.id, 'student_id': r.student_id,
        'student_name': stu.name if stu else '',
        'student_no': stu.student_no if stu else '',
        'class_name': cls_name,
        'record_date': r.record_date, 'assessment_date': r.record_date,   # alias for frontend
        'location': r.location, 'topic': r.topic, 'summary': r.summary,
        'notes': r.summary,                                                 # alias
        'emotion_tags': r.emotion_tags, 'follow_up_plan': r.follow_up_plan,
        'next_follow_date': r.next_follow_date, 'next_follow_up': r.next_follow_date,  # alias
        'attention_level': getattr(r, 'attention_level', '') or '',
        'counseling_count': getattr(r, 'counseling_count', 0) or 0,
        'reminded': bool(getattr(r, 'reminded', False)),
        'reminded_at': getattr(r, 'reminded_at', None).isoformat() if getattr(r, 'reminded_at', None) else None,
    }


def _psy_normalize_input(data: dict) -> dict:
    """接收前端字段 -> 转成 model 字段"""
    d = dict(data)
    if 'assessment_date' in d and 'record_date' not in d:
        d['record_date'] = d.pop('assessment_date')
    if 'next_follow_up' in d and 'next_follow_date' not in d:
        d['next_follow_date'] = d.pop('next_follow_up')
    if 'notes' in d and 'summary' not in d:
        d['summary'] = d.pop('notes')
    # 只保留 model 有的字段
    allowed = {'student_id','record_date','location','topic','summary','emotion_tags','follow_up_plan','next_follow_date','attention_level','counseling_count'}
    return {k: v for k, v in d.items() if k in allowed}


@router.get('/psychology')
def list_psychology(
    student_id: Optional[int] = Query(None),
    attention_level: Optional[str] = Query(None),
    search: str = Query('', description='搜索学号/姓名/主题/备注'),
    sort_by: str = Query('record_date', description='排序字段'),
    order: str = Query('desc', description='asc/desc'),
    db: Session = Depends(get_db)
):
    """心理档案列表 (v3j-B-b03 · 支持 search + sort_by + order)"""
    q = db.query(PsychologyRecord).outerjoin(Student, PsychologyRecord.student_id == Student.id)
    if student_id:
        q = q.filter(PsychologyRecord.student_id == student_id)
    if attention_level:
        q = q.filter(PsychologyRecord.attention_level == attention_level)
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Student.name.ilike(pattern),
                Student.student_no.ilike(pattern),
                PsychologyRecord.topic.ilike(pattern),
                PsychologyRecord.summary.ilike(pattern),
                PsychologyRecord.attention_level.ilike(pattern),
            )
        )
    SORT_WHITELIST = {
        'record_date': PsychologyRecord.record_date,
        'assessment_date': PsychologyRecord.record_date,
        'attention_level': PsychologyRecord.attention_level,
        'counseling_count': PsychologyRecord.counseling_count,
        'next_follow_date': PsychologyRecord.next_follow_date,
        'next_follow_up': PsychologyRecord.next_follow_date,
        'student_name': Student.name,
        'student_no': Student.student_no,
    }
    col = SORT_WHITELIST.get(sort_by, PsychologyRecord.record_date)
    if (order or 'desc').lower() == 'asc':
        q = q.order_by(col.asc())
    else:
        q = q.order_by(col.desc())
    items = q.all()
    return [_psy_dict(r, db) for r in items]


def _psy_query_with_filters(db: Session, search: str = '', student_id: Optional[int] = None, attention_level: Optional[str] = None):
    q = db.query(PsychologyRecord).outerjoin(Student, PsychologyRecord.student_id == Student.id)
    if student_id:
        q = q.filter(PsychologyRecord.student_id == student_id)
    if attention_level:
        q = q.filter(PsychologyRecord.attention_level == attention_level)
    if search:
        pattern = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Student.name.ilike(pattern),
                Student.student_no.ilike(pattern),
                PsychologyRecord.topic.ilike(pattern),
                PsychologyRecord.summary.ilike(pattern),
                PsychologyRecord.attention_level.ilike(pattern),
            )
        )
    return q


def _psy_export_workbook(rows, db):
    from openpyxl import Workbook
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = '心理档案'
    ws.append(['学号', '姓名', '班级', '关注等级', '测评日期', '咨询次数', '下次跟进', '主题', '备注'])
    for r in rows:
        d = _psy_dict(r, db)
        ws.append([
            d['student_no'], d['student_name'], d['class_name'],
            d['attention_level'], d['record_date'],
            d['counseling_count'], d['next_follow_date'],
            d.get('topic', ''), d.get('summary', ''),
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.post('/psychology/export')
def export_psychology_by_ids(payload: dict = Body(...), db: Session = Depends(get_db)):
    """按 ID 列表批量导出心理档案 Excel (v3j-B-b03)"""
    from fastapi.responses import StreamingResponse
    ids = payload.get('ids') or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, '请传入非空的 ids 列表')
    rows = db.query(PsychologyRecord).filter(PsychologyRecord.id.in_(ids)).all()
    output = _psy_export_workbook(rows, db)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=psychology_selected.xlsx'}
    )


@router.get('/psychology/export/all')
def export_psychology_all(
    search: str = Query(''),
    student_id: Optional[int] = Query(None),
    attention_level: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """按当前搜索条件导出全部心理档案 Excel (v3j-B-b03)"""
    from fastapi.responses import StreamingResponse
    q = _psy_query_with_filters(db, search, student_id, attention_level)
    rows = q.order_by(PsychologyRecord.record_date.desc()).all()
    output = _psy_export_workbook(rows, db)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=psychology_all.xlsx'}
    )


@router.get('/psychology/reminders')
def psychology_reminders(db: Session = Depends(get_db)):
    """下次跟进到期提醒"""
    from datetime import datetime, timedelta
    today = datetime.now().strftime('%Y-%m-%d')
    items = db.query(PsychologyRecord).filter(
        PsychologyRecord.next_follow_date != '',
        PsychologyRecord.next_follow_date <= today,
        (PsychologyRecord.reminded == False) | (PsychologyRecord.reminded.is_(None))
    ).all()
    result = []
    for r in items:
        stu = db.get(Student, r.student_id)
        result.append({
            'id': r.id, 'student_name': stu.name if stu else '',
            'topic': r.topic, 'next_follow_date': r.next_follow_date,
        })
    return result


# ===== 心理关怀 统计图表 =====
@router.get('/psychology/chart-data')
def psychology_chart_data(db: Session = Depends(get_db)):
    """心理关怀统计图表数据"""
    # 1. level_distribution: GROUP BY attention_level, COUNT(*)
    level_rows = db.query(
        PsychologyRecord.attention_level, func.count(PsychologyRecord.id)
    ).group_by(PsychologyRecord.attention_level).all()
    level_distribution = [{'level': lv or '', 'count': c} for lv, c in level_rows]

    # 2. monthly_trend: 按 record_date 的 YYYY-MM 分组，最近12个月
    today = datetime.now()
    months = [(today - timedelta(days=30 * i)).strftime('%Y-%m') for i in range(11, -1, -1)]
    month_rows = db.query(
        func.substr(PsychologyRecord.record_date, 1, 7).label('month'),
        func.count(PsychologyRecord.id)
    ).filter(
        PsychologyRecord.record_date != '', PsychologyRecord.record_date.isnot(None)
    ).group_by(
        func.substr(PsychologyRecord.record_date, 1, 7)
    ).all()
    month_map = {m: c for m, c in month_rows if m}
    monthly_trend = [{'month': m, 'count': month_map.get(m, 0)} for m in months]

    # 3. emotion_tags_distribution: 解析 JSON array 字符串，展平统计
    tag_counter = Counter()
    all_tags_rows = db.query(PsychologyRecord.emotion_tags).filter(
        PsychologyRecord.emotion_tags != '', PsychologyRecord.emotion_tags.isnot(None)
    ).all()
    for (tags_str,) in all_tags_rows:
        try:
            tags = _json.loads(tags_str)
            if isinstance(tags, list):
                tag_counter.update(tags)
        except (ValueError, TypeError):
            pass
    emotion_tags_distribution = [{'tag': t, 'count': c} for t, c in tag_counter.most_common()]

    # 4. top_students: 按 student_id 分组，取 TOP10
    top_rows = db.query(
        PsychologyRecord.student_id, func.count(PsychologyRecord.id).label('cnt')
    ).group_by(PsychologyRecord.student_id).order_by(func.count(PsychologyRecord.id).desc()).limit(10).all()
    top_students = []
    for sid, cnt in top_rows:
        stu = db.get(Student, sid)
        top_students.append({
            'student_name': stu.name if stu else '',
            'student_no': stu.student_no if stu else '',
            'count': cnt,
        })

    return {
        'level_distribution': level_distribution,
        'monthly_trend': monthly_trend,
        'emotion_tags_distribution': emotion_tags_distribution,
        'top_students': top_students,
    }


# ===== 家校沟通 =====
def _fam_dict(c, db):
    from models import ClassModel
    stu = db.get(Student, c.student_id)
    cls_name = ''
    if stu and stu.class_id:
        cls = db.get(ClassModel, stu.class_id)
        cls_name = cls.class_name if cls else ''
    # 从 parent_name '父亲XX' / '母亲XX' 抽出关系
    _rel = ''
    for k in ('父亲','母亲','监护人','祖父','祖母','外祖父','外祖母','兄弟','姐妹'):
        if c.parent_name and c.parent_name.startswith(k):
            _rel = k
            break
    return {
        'id': c.id, 'student_id': c.student_id,
        'student_name': stu.name if stu else '',
        'student_no': stu.student_no if stu else '',
        'class_name': cls_name,
        'contact_date': c.contact_date,
        'parent_name': c.parent_name, 'contact_name': c.parent_name,   # alias
        'contact_method': c.contact_method, 'contact_type': c.contact_method,  # alias
        'relationship': _rel,
        'topic': c.topic,
        'conclusion': c.conclusion, 'content': c.conclusion,   # alias
        'attachment': c.attachment,
    }


def _fam_normalize_input(data: dict) -> dict:
    d = dict(data)
    if 'contact_name' in d and 'parent_name' not in d:
        d['parent_name'] = d.pop('contact_name')
    if 'contact_type' in d and 'contact_method' not in d:
        d['contact_method'] = d.pop('contact_type')
    if 'content' in d and 'conclusion' not in d:
        d['conclusion'] = d.pop('content')
    # 如果传了 relationship 但 parent_name 没前缀，拼进去
    rel = d.pop('relationship', None)
    if rel and d.get('parent_name') and not any(d['parent_name'].startswith(k) for k in ('父亲','母亲','监护人','祖父','祖母','外祖父','外祖母','兄弟','姐妹')):
        d['parent_name'] = f"{rel}{d['parent_name']}"
    allowed = {'student_id','contact_date','parent_name','contact_method','topic','conclusion','attachment'}
    return {k: v for k, v in d.items() if k in allowed}


@router.get('/psychology/{record_id}')
def get_psychology(record_id: int, db: Session = Depends(get_db)):
    r = db.get(PsychologyRecord, record_id)
    if not r:
        raise HTTPException(404, '记录不存在')
    return _psy_dict(r, db)


@router.post('/psychology')
def create_psychology(data: dict, db: Session = Depends(get_db)):
    r = PsychologyRecord(**_psy_normalize_input(data))
    db.add(r)
    db.commit()
    db.refresh(r)
    return {'id': r.id}


@router.put('/psychology/{rid}')
def update_psychology(rid: int, data: dict, db: Session = Depends(get_db)):
    r = db.get(PsychologyRecord, rid)
    if not r:
        raise HTTPException(404, '记录不存在')
    for k, v in _psy_normalize_input(data).items():
        if hasattr(r, k):
            setattr(r, k, v)
    db.commit()
    return {'ok': True}


@router.delete('/psychology/{rid}')
def delete_psychology(rid: int, db: Session = Depends(get_db)):
    r = db.get(PsychologyRecord, rid)
    if r:
        db.delete(r)
        db.commit()
    return {'ok': True}




@router.get('/family-contacts')
def list_family_contacts(student_id: Optional[int] = None, contact_type: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(FamilyContact)
    if student_id:
        q = q.filter(FamilyContact.student_id == student_id)
    if contact_type:
        q = q.filter(FamilyContact.contact_method == contact_type)
    items = q.order_by(FamilyContact.created_at.desc()).all()
    return [_fam_dict(c, db) for c in items]


@router.get('/family-contacts/{contact_id}')
def get_family_contact(contact_id: int, db: Session = Depends(get_db)):
    c = db.get(FamilyContact, contact_id)
    if not c:
        raise HTTPException(404, '记录不存在')
    return _fam_dict(c, db)


@router.post('/family-contacts')
def create_family_contact(data: dict, db: Session = Depends(get_db)):
    c = FamilyContact(**_fam_normalize_input(data))
    db.add(c)
    db.commit()
    db.refresh(c)
    return {'id': c.id}


@router.put('/family-contacts/{cid}')
def update_family_contact(cid: int, data: dict, db: Session = Depends(get_db)):
    c = db.get(FamilyContact, cid)
    if not c:
        raise HTTPException(404, '记录不存在')
    for k, v in _fam_normalize_input(data).items():
        if hasattr(c, k):
            setattr(c, k, v)
    db.commit()
    return {'ok': True}


@router.delete('/family-contacts/{cid}')
def delete_family_contact(cid: int, db: Session = Depends(get_db)):
    c = db.get(FamilyContact, cid)
    if c:
        db.delete(c)
        db.commit()
    return {'ok': True}



# ===== v3j-D 补丁2: 心理关怀已提醒 toggle + 批量提醒 =====
@router.patch('/psychology/{rec_id}/toggle-reminded')
def toggle_psy_reminded(rec_id: int, db: Session = Depends(get_db)):
    """切换心理档案的已提醒状态"""
    from datetime import datetime as _dt
    r = db.get(PsychologyRecord, rec_id)
    if not r:
        raise HTTPException(404, '记录不存在')
    r.reminded = not bool(getattr(r, 'reminded', False))
    r.reminded_at = _dt.now() if r.reminded else None
    db.commit()
    db.refresh(r)
    return {'id': r.id, 'reminded': bool(r.reminded), 'reminded_at': r.reminded_at.isoformat() if r.reminded_at else None}


@router.post('/psychology/batch-mark-reminded')
def batch_mark_psy_reminded(payload: dict = Body(...), db: Session = Depends(get_db)):
    """批量标记心理档案已提醒。payload: {ids: [...], reminded: true/false}"""
    from datetime import datetime as _dt
    ids = payload.get('ids') or []
    reminded = bool(payload.get('reminded', True))
    if not ids:
        return {'updated': 0}
    now = _dt.now() if reminded else None
    updated = db.query(PsychologyRecord).filter(PsychologyRecord.id.in_(ids)).update(
        {PsychologyRecord.reminded: reminded, PsychologyRecord.reminded_at: now},
        synchronize_session=False
    )
    db.commit()
    return {'updated': updated, 'reminded': reminded}
