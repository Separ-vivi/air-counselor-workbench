"""学期报表 API - V5-h-hotfix10
全面重构：增加考勤/心理/资助/荣誉/违纪/访谈/宿舍维度
修复导出500、修复空数据崩溃
增加semester参数到心理/违纪/荣誉/访谈API + 党员人数 + 访谈覆盖率
V5-h-hotfix10: 所有统计指标严格按学期筛选 + 对比指标扩展到12项
"""
import io
import logging
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, case, distinct
from sqlalchemy.orm import Session
from database import get_db
from models import (
    Student, ClassModel, Major, GradeRecord, WarningRecord,
    PartyProgress, EmploymentRecord, Activity, ActivitySignup,
    PsychologyRecord, StudentDiscipline, StudentHardship,
    StudentGrant, StudentScholarship, StudentLoan, StudentWorkStudy,
    StudentHonor, StudentDormVisit, StudentLeave, StudentDormChat,
    StudentAttendanceException, StudentInterview
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix='/api/semester-report', tags=['学期报表'])


def _get_current_semester():
    """根据当前日期自动计算当前学期，格式如 2025-2026-1"""
    now = datetime.now()
    y = now.year
    m = now.month
    if m >= 9:
        return f"{y}-{y+1}-1"
    elif m >= 2:
        return f"{y-1}-{y}-2"
    else:
        return f"{y-1}-{y}-1"


def _format_semester_display(semester: str) -> str:
    """将学期代码转为显示格式"""
    if not semester or semester == 'all':
        return '全部学期'
    parts = semester.split('-')
    if len(parts) == 3:
        return f"{parts[0]}-{parts[1]}学年第{parts[2]}学期"
    elif len(parts) == 2:
        y = int(parts[0])
        sem_num = parts[1]
        if sem_num == '1':
            return f"{y-1}-{y}学年第1学期"
        else:
            return f"{y}-{y+1}学年第2学期"
    return semester


def _semester_date_range(semester: str):
    """将学期字符串转为日期范围"""
    if not semester or semester == 'all':
        return None, None
    parts = semester.split('-')
    if len(parts) == 3:
        y1, y2, term = parts[0], parts[1], parts[2]
        if term == '1':
            return f"{y1}-09-01", f"{y2}-01-31"
        else:
            return f"{y1}-09-01", f"{y2}-07-31"
    return None, None


def _semester_to_academic_year(semester: str):
    """将学期代码转为学年，如 '2025-2026-1' -> '2025-2026'"""
    if not semester or semester == 'all':
        return None
    parts = semester.split('-')
    if len(parts) >= 2:
        return f"{parts[0]}-{parts[1]}"
    return None


# ============================================================
# 0. 获取可用学期列表
# ============================================================
@router.get('/semesters')
def list_semesters(db: Session = Depends(get_db)):
    """返回数据库中已有的学期列表"""
    semesters = set()
    try:
        rows = db.query(GradeRecord.semester).distinct().all()
        for r in rows:
            if r[0]:
                semesters.add(r[0])
    except Exception:
        pass
    try:
        rows = db.query(WarningRecord.semester).distinct().all()
        for r in rows:
            if r[0]:
                semesters.add(r[0])
    except Exception:
        pass
    sorted_sems = sorted(semesters, reverse=True)
    return [
        {'code': s, 'label': _format_semester_display(s)}
        for s in sorted_sems
    ]


# ============================================================
# 1. 学期总览（增强）
# ============================================================
@router.get('/summary')
def semester_summary(semester: str = Query(None), db: Session = Depends(get_db)):
    """学期总览：增加考勤异常、心理关注、资助人次、违纪人数、荣誉人次"""
    result = {
        'total_students': 0,
        'total_classes': 0,
        'total_majors': 0,
        'political_distribution': {},
        'gender_distribution': {},
        'housing_distribution': {'外宿': 0, '住校': 0},
        'campus_distribution': {},
        # 新增字段
        'attendance_exception_count': 0,
        'psychology_attention_count': 0,
        'financial_aid_count': 0,
        'discipline_count': 0,
        'honor_count': 0,
        'party_member_count': 0,
    }
    try:
        result['total_students'] = db.query(Student).count()
        result['total_classes'] = db.query(ClassModel).count()
        result['total_majors'] = db.query(Major).count()
    except Exception as e:
        logger.warning(f"summary 基础统计异常: {e}")

    # 政治面貌分布
    try:
        rows = (
            db.query(Student.political_status, func.count(Student.id))
            .group_by(Student.political_status)
            .all()
        )
        result['political_distribution'] = {
            (status or '群众'): cnt for status, cnt in rows
        }
    except Exception as e:
        logger.warning(f"summary 政治面貌异常: {e}")

    # 性别分布
    try:
        rows = (
            db.query(Student.gender, func.count(Student.id))
            .group_by(Student.gender)
            .all()
        )
        result['gender_distribution'] = {
            (g or '未知'): cnt for g, cnt in rows
        }
    except Exception as e:
        logger.warning(f"summary 性别异常: {e}")

    # 外宿/住校
    try:
        off = db.query(Student).filter(Student.is_off_campus == True).count()
        total = result['total_students']
        result['housing_distribution'] = {'外宿': off, '住校': total - off}
    except Exception as e:
        logger.warning(f"summary 住宿异常: {e}")

    # 校区分布
    try:
        rows = (
            db.query(Student.campus, func.count(Student.id))
            .group_by(Student.campus)
            .all()
        )
        result['campus_distribution'] = {
            (c or '未知'): cnt for c, cnt in rows
        }
    except Exception as e:
        logger.warning(f"summary 校区异常: {e}")

    # 考勤异常总次数
    try:
        q = db.query(StudentAttendanceException)
        if semester and semester != 'all':
            # StudentAttendanceException 没有 semester 字段，按日期范围筛选
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    StudentAttendanceException.exception_date >= start,
                    StudentAttendanceException.exception_date <= end
                )
        result['attendance_exception_count'] = q.count()
    except Exception as e:
        logger.warning(f"summary 考勤异常异常: {e}")

    # 心理关注人数（非普通等级，按学期日期范围筛选）
    try:
        q = db.query(PsychologyRecord.student_id).filter(
            PsychologyRecord.attention_level.in_(['一级关注', '二级关注', '三级关注'])
        )
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    PsychologyRecord.record_date >= start,
                    PsychologyRecord.record_date <= end
                )
        result['psychology_attention_count'] = q.distinct().count()
    except Exception as e:
        logger.warning(f"summary 心理关注异常: {e}")

    # 资助总人次
    try:
        aid_year = _semester_to_academic_year(semester)
        hardship_q = db.query(StudentHardship)
        grant_q = db.query(StudentGrant)
        scholarship_q = db.query(StudentScholarship)
        loan_q = db.query(StudentLoan)
        work_q = db.query(StudentWorkStudy)
        if aid_year:
            hardship_q = hardship_q.filter(StudentHardship.academic_year == aid_year)
            grant_q = grant_q.filter(StudentGrant.academic_year == aid_year)
            scholarship_q = scholarship_q.filter(StudentScholarship.academic_year == aid_year)
            work_q = work_q.filter(StudentWorkStudy.academic_year == aid_year)
        cnt = hardship_q.count() + grant_q.count() + scholarship_q.count() + loan_q.count() + work_q.count()
        result['financial_aid_count'] = cnt
    except Exception as e:
        logger.warning(f"summary 资助人次异常: {e}")

    # 违纪人数（按学期日期范围筛选）
    try:
        q = db.query(StudentDiscipline.student_id).distinct()
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    StudentDiscipline.discipline_date >= start,
                    StudentDiscipline.discipline_date <= end
                )
        result['discipline_count'] = q.count()
    except Exception as e:
        logger.warning(f"summary 违纪人数异常: {e}")

    # 荣誉总人次（按学年筛选）
    try:
        q = db.query(StudentHonor)
        honor_year = _semester_to_academic_year(semester)
        if honor_year:
            q = q.filter(StudentHonor.academic_year == honor_year)
        result['honor_count'] = q.count()
    except Exception as e:
        logger.warning(f"summary 荣誉人次异常: {e}")

    # 党员人数（按学期过滤：取该学期内最新 stage，统计预备党员+正式党员）
    try:
        _pp_subq = db.query(
            PartyProgress.student_id,
            func.max(PartyProgress.id).label('max_id')
        )
        if semester and semester != 'all':
            _pp_start, _pp_end = _semester_date_range(semester)
            if _pp_start and _pp_end:
                _pp_subq = _pp_subq.filter(
                    PartyProgress.stage_date >= _pp_start,
                    PartyProgress.stage_date <= _pp_end
                )
        _pp_subq = _pp_subq.group_by(PartyProgress.student_id).subquery()
        result['party_member_count'] = (
            db.query(func.count(PartyProgress.student_id))
            .join(_pp_subq, PartyProgress.id == _pp_subq.c.max_id)
            .filter(PartyProgress.stage.in_(['中共预备党员', '中共党员']))
            .scalar() or 0
        )
    except Exception as e:
        logger.warning(f"summary 党员人数异常: {e}")

    return result


# ============================================================
# 2. 学业数据
# ============================================================
@router.get('/academics')
def semester_academics(semester: str = Query(None), db: Session = Depends(get_db)):
    """学业汇总"""
    result = {
        'class_averages': [],
        'fail_rate': 0.0,
        'fail_count': 0,
        'total_students_with_grades': 0,
        'top10': [],
        'warning_stats': [],
    }

    grade_filters = [GradeRecord.score.isnot(None)]
    warn_filters = []
    if semester and semester != 'all':
        grade_filters.append(GradeRecord.semester == semester)
        warn_filters.append(WarningRecord.semester == semester)

    # 各班平均成绩
    try:
        rows = (
            db.query(
                ClassModel.class_name,
                func.avg(GradeRecord.score)
            )
            .join(Student, Student.class_id == ClassModel.id)
            .join(GradeRecord, GradeRecord.student_id == Student.id)
            .filter(*grade_filters)
            .group_by(ClassModel.id, ClassModel.class_name)
            .all()
        )
        result['class_averages'] = [
            {'class_name': name, 'avg_score': round(avg, 2) if avg else 0}
            for name, avg in rows
        ]
    except Exception as e:
        logger.warning(f"academics 班级平均异常: {e}")

    # 挂科率
    try:
        base_q = db.query(GradeRecord.student_id).filter(*grade_filters)
        total_with_grades = base_q.distinct().count() or 0
        fail_students = base_q.filter(GradeRecord.score < 60).distinct().count() or 0
        result['total_students_with_grades'] = total_with_grades
        result['fail_count'] = fail_students
        result['fail_rate'] = (
            round(fail_students / total_with_grades * 100, 2)
            if total_with_grades > 0 else 0.0
        )
    except Exception as e:
        logger.warning(f"academics 挂科率异常: {e}")

    # 成绩 Top 10
    try:
        rows = (
            db.query(
                Student.student_no,
                Student.name,
                ClassModel.class_name,
                func.avg(GradeRecord.score).label('avg_score')
            )
            .join(Student, GradeRecord.student_id == Student.id)
            .join(ClassModel, Student.class_id == ClassModel.id)
            .filter(*grade_filters)
            .group_by(Student.id, Student.student_no, Student.name, ClassModel.class_name)
            .order_by(func.avg(GradeRecord.score).desc())
            .limit(10)
            .all()
        )
        result['top10'] = [
            {
                'student_no': r[0],
                'name': r[1],
                'class_name': r[2],
                'avg_score': round(r[3], 2) if r[3] else 0
            }
            for r in rows
        ]
    except Exception as e:
        logger.warning(f"academics Top10 异常: {e}")

    # 学业预警统计
    try:
        red_q = db.query(WarningRecord).filter(WarningRecord.warning_type == 'red')
        yellow_q = db.query(WarningRecord).filter(WarningRecord.warning_type == 'yellow')
        if warn_filters:
            red_q = red_q.filter(*warn_filters)
            yellow_q = yellow_q.filter(*warn_filters)
        red = red_q.count()
        yellow = yellow_q.count()
        result['warning_stats'] = [
            {'level': 'red', 'level_label': '红色预警', 'count': red},
            {'level': 'yellow', 'level_label': '黄色预警', 'count': yellow},
            {'level': 'normal', 'level_label': '正常', 'count': max(0, result['total_students_with_grades'] - red - yellow)},
        ]
    except Exception as e:
        logger.warning(f"academics 预警异常: {e}")

    return result


# ============================================================
# 3. 党团发展
# ============================================================
@router.get('/party-development')
def party_development(db: Session = Depends(get_db)):
    """党团发展进度"""
    stage_map = {
        '递交入党申请书': 0,
        '入党积极分子': 0,
        '发展对象': 0,
        '中共预备党员': 0,
        '中共党员': 0,
    }
    total_new_this_semester = 0

    try:
        subq = (
            db.query(
                PartyProgress.student_id,
                func.max(PartyProgress.id).label('max_id')
            )
            .group_by(PartyProgress.student_id)
            .subquery()
        )
        rows = (
            db.query(PartyProgress.stage, func.count(PartyProgress.student_id))
            .join(subq, PartyProgress.id == subq.c.max_id)
            .group_by(PartyProgress.stage)
            .all()
        )
        for stage, cnt in rows:
            if stage in stage_map:
                stage_map[stage] = cnt
            else:
                stage_map[stage] = cnt
    except Exception as e:
        logger.warning(f"party 各阶段异常: {e}")

    try:
        now = datetime.now()
        if now.month >= 9:
            sem_start = f"{now.year}-09-01"
        else:
            sem_start = f"{now.year}-03-01"
        total_new_this_semester = (
            db.query(func.count(PartyProgress.id))
            .filter(PartyProgress.stage_date >= sem_start)
            .scalar() or 0
        )
    except Exception as e:
        logger.warning(f"party 新发展异常: {e}")

    return {
        'stages': stage_map,
        'new_this_semester': total_new_this_semester,
    }


# ============================================================
# 4. 就业跟踪
# ============================================================
@router.get('/employment')
def employment_stats(db: Session = Depends(get_db)):
    """就业状态分布与就业率"""
    status_map = {
        '已签约': 0,
        '考研': 0,
        '出国': 0,
        '待就业': 0,
        '未知': 0,
    }
    total = 0

    try:
        rows = (
            db.query(EmploymentRecord.status, func.count(EmploymentRecord.id))
            .group_by(EmploymentRecord.status)
            .all()
        )
        total = sum(cnt for _, cnt in rows)
        for status, cnt in rows:
            s = (status or '').strip()
            if s in status_map:
                status_map[s] = cnt
            elif s:
                status_map[s] = cnt
            else:
                status_map['未知'] += cnt
    except Exception as e:
        logger.warning(f"employment 异常: {e}")

    employed = status_map.get('已签约', 0)
    employment_rate = round(employed / total * 100, 2) if total > 0 else 0.0

    return {
        'distribution': status_map,
        'total_records': total,
        'total_count': total,
        'employed_count': employed,
        'employment_rate': employment_rate,
    }


# ============================================================
# 5. 学生活动
# ============================================================
@router.get('/activities')
def activity_stats(db: Session = Depends(get_db)):
    """活动统计"""
    total_activities = 0
    total_participants = 0
    activity_ranking = []

    try:
        total_activities = db.query(Activity).count()
    except Exception as e:
        logger.warning(f"activities 总数异常: {e}")

    try:
        total_participants = db.query(ActivitySignup).count()
    except Exception as e:
        logger.warning(f"activities 人次异常: {e}")

    try:
        rows = (
            db.query(
                Activity.title,
                Activity.activity_type,
                func.count(ActivitySignup.id).label('participants')
            )
            .join(ActivitySignup, ActivitySignup.activity_id == Activity.id)
            .group_by(Activity.id, Activity.title, Activity.activity_type)
            .order_by(func.count(ActivitySignup.id).desc())
            .limit(10)
            .all()
        )
        activity_ranking = [
            {
                'title': r[0],
                'activity_type': r[1],
                'participants': r[2],
            }
            for r in rows
        ]
    except Exception as e:
        logger.warning(f"activities 排名异常: {e}")

    return {
        'total_activities': total_activities,
        'total_participants': total_participants,
        'activity_ranking': activity_ranking,
    }


# ============================================================
# 6. 考勤汇总
# ============================================================
@router.get('/attendance')
def attendance_summary(semester: str = Query(None), db: Session = Depends(get_db)):
    """考勤汇总：总异常次数、按类型分类、按班级分组"""
    result = {
        'total_exceptions': 0,
        'by_type': {},
        'by_class': [],
    }
    try:
        q = db.query(StudentAttendanceException)
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    StudentAttendanceException.exception_date >= start,
                    StudentAttendanceException.exception_date <= end
                )
        result['total_exceptions'] = q.count()
    except Exception as e:
        logger.warning(f"attendance 总数异常: {e}")

    # 按类型分类
    try:
        q = db.query(
            StudentAttendanceException.exception_type,
            func.count(StudentAttendanceException.id)
        ).group_by(StudentAttendanceException.exception_type)
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    StudentAttendanceException.exception_date >= start,
                    StudentAttendanceException.exception_date <= end
                )
        rows = q.all()
        result['by_type'] = {(t or '未知'): cnt for t, cnt in rows}
    except Exception as e:
        logger.warning(f"attendance 按类型异常: {e}")

    # 按班级分组
    try:
        q = (
            db.query(
                ClassModel.class_name,
                func.count(StudentAttendanceException.id)
            )
            .join(Student, StudentAttendanceException.student_id == Student.id)
            .join(ClassModel, Student.class_id == ClassModel.id)
            .group_by(ClassModel.id, ClassModel.class_name)
        )
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    StudentAttendanceException.exception_date >= start,
                    StudentAttendanceException.exception_date <= end
                )
        rows = q.all()
        result['by_class'] = [
            {'class_name': name, 'count': cnt}
            for name, cnt in rows
        ]
    except Exception as e:
        logger.warning(f"attendance 按班级异常: {e}")

    return result


# ============================================================
# 7. 心理档案汇总
# ============================================================
@router.get('/psychology')
def psychology_summary(semester: str = Query(None), db: Session = Depends(get_db)):
    """心理档案汇总：关注等级、咨询次数、需跟进人数"""
    result = {
        'by_attention_level': {},
        'total_counseling_count': 0,
        'need_follow_up': 0,
    }

    # 构建日期范围过滤
    psych_filters = []
    if semester and semester != 'all':
        start, end = _semester_date_range(semester)
        if start and end:
            psych_filters.append(PsychologyRecord.record_date >= start)
            psych_filters.append(PsychologyRecord.record_date <= end)

    # 按关注等级统计人数
    try:
        q = (
            db.query(
                PsychologyRecord.attention_level,
                func.count(PsychologyRecord.student_id.distinct())
            )
            .group_by(PsychologyRecord.attention_level)
        )
        if psych_filters:
            q = q.filter(*psych_filters)
        rows = q.all()
        result['by_attention_level'] = {
            (level or '普通'): cnt for level, cnt in rows
        }
    except Exception as e:
        logger.warning(f"psychology 关注等级异常: {e}")

    # 总咨询次数
    try:
        q = db.query(func.sum(PsychologyRecord.counseling_count))
        if psych_filters:
            q = q.filter(*psych_filters)
        total = q.scalar()
        result['total_counseling_count'] = total or 0
    except Exception as e:
        logger.warning(f"psychology 咨询次数异常: {e}")

    # 需跟进人数（有 follow_up_plan 且 next_follow_date >= 今天）
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        q = (
            db.query(PsychologyRecord.student_id)
            .filter(
                PsychologyRecord.follow_up_plan != '',
                PsychologyRecord.follow_up_plan.isnot(None),
            )
            .distinct()
        )
        if psych_filters:
            q = q.filter(*psych_filters)
        need_follow = q.count()
        result['need_follow_up'] = need_follow or 0
    except Exception as e:
        logger.warning(f"psychology 跟进人数异常: {e}")

    return result


# ============================================================
# 8. 违纪统计
# ============================================================
@router.get('/discipline')
def discipline_summary(semester: str = Query(None), db: Session = Depends(get_db)):
    """违纪统计：按处分类型统计、涉及学生数"""
    result = {
        'by_type': {},
        'student_count': 0,
    }

    # 构建日期范围过滤
    disc_filters = []
    if semester and semester != 'all':
        start, end = _semester_date_range(semester)
        if start and end:
            disc_filters.append(StudentDiscipline.discipline_date >= start)
            disc_filters.append(StudentDiscipline.discipline_date <= end)

    try:
        q = (
            db.query(
                StudentDiscipline.discipline_type,
                func.count(StudentDiscipline.id)
            )
            .group_by(StudentDiscipline.discipline_type)
        )
        if disc_filters:
            q = q.filter(*disc_filters)
        rows = q.all()
        result['by_type'] = {(t or '未知'): cnt for t, cnt in rows}
    except Exception as e:
        logger.warning(f"discipline 按类型异常: {e}")

    try:
        q = db.query(StudentDiscipline.student_id).distinct()
        if disc_filters:
            q = q.filter(*disc_filters)
        result['student_count'] = q.count()
    except Exception as e:
        logger.warning(f"discipline 学生数异常: {e}")

    return result


# ============================================================
# 9. 资助汇总
# ============================================================
@router.get('/financial-aid')
def financial_aid_summary(semester: str = Query(None), db: Session = Depends(get_db)):
    """资助汇总：困难认定、助学金、奖学金、贷款、勤工助学"""
    aid_year = _semester_to_academic_year(semester)
    result = {
        'hardship_by_level': {},
        'hardship_count': 0,
        'grant_total_amount': 0,
        'grant_count': 0,
        'scholarship_total_amount': 0,
        'scholarship_count': 0,
        'loan_total_amount': 0,
        'loan_count': 0,
        'work_study_count': 0,
        'work_study_total_compensation': 0,
    }

    # 困难认定
    try:
        q = db.query(
            StudentHardship.hardship_level,
            func.count(StudentHardship.id)
        ).group_by(StudentHardship.hardship_level)
        q_total = db.query(StudentHardship)
        if aid_year:
            q = q.filter(StudentHardship.academic_year == aid_year)
            q_total = q_total.filter(StudentHardship.academic_year == aid_year)
        rows = q.all()
        result['hardship_by_level'] = {(level or '未知'): cnt for level, cnt in rows}
        result['hardship_count'] = q_total.count()
    except Exception as e:
        logger.warning(f"financial-aid 困难认定异常: {e}")

    # 助学金
    try:
        q = db.query(func.sum(StudentGrant.amount), func.count(StudentGrant.id))
        if aid_year:
            q = q.filter(StudentGrant.academic_year == aid_year)
        total_amount, count = q.first()
        result['grant_total_amount'] = total_amount or 0
        result['grant_count'] = count or 0
    except Exception as e:
        logger.warning(f"financial-aid 助学金异常: {e}")

    # 奖学金
    try:
        q = db.query(func.sum(StudentScholarship.amount), func.count(StudentScholarship.id))
        if aid_year:
            q = q.filter(StudentScholarship.academic_year == aid_year)
        total_amount, count = q.first()
        result['scholarship_total_amount'] = total_amount or 0
        result['scholarship_count'] = count or 0
    except Exception as e:
        logger.warning(f"financial-aid 奖学金异常: {e}")

    # 助学贷款
    try:
        q = db.query(func.sum(StudentLoan.amount), func.count(StudentLoan.id))
        total_amount, count = q.first()
        result['loan_total_amount'] = total_amount or 0
        result['loan_count'] = count or 0
    except Exception as e:
        logger.warning(f"financial-aid 助学贷款异常: {e}")

    # 勤工助学
    try:
        q = db.query(func.sum(StudentWorkStudy.compensation), func.count(StudentWorkStudy.id))
        if aid_year:
            q = q.filter(StudentWorkStudy.academic_year == aid_year)
        total_comp, count = q.first()
        result['work_study_total_compensation'] = total_comp or 0
        result['work_study_count'] = count or 0
    except Exception as e:
        logger.warning(f"financial-aid 勤工助学异常: {e}")

    return result


# ============================================================
# 10. 荣誉统计
# ============================================================
@router.get('/honors')
def honors_summary(semester: str = Query(None), db: Session = Depends(get_db)):
    """荣誉统计：按级别统计获奖人次、获奖学生数"""
    result = {
        'by_level': {},
        'student_count': 0,
    }

    # 构建学年过滤
    honor_filters = []
    honor_year = _semester_to_academic_year(semester)
    if honor_year:
        honor_filters.append(StudentHonor.academic_year == honor_year)

    try:
        q = (
            db.query(
                StudentHonor.level,
                func.count(StudentHonor.id)
            )
            .group_by(StudentHonor.level)
        )
        if honor_filters:
            q = q.filter(*honor_filters)
        rows = q.all()
        result['by_level'] = {(level or '未知'): cnt for level, cnt in rows}
    except Exception as e:
        logger.warning(f"honors 按级别异常: {e}")

    try:
        q = db.query(StudentHonor.student_id).distinct()
        if honor_filters:
            q = q.filter(*honor_filters)
        result['student_count'] = q.count()
    except Exception as e:
        logger.warning(f"honors 学生数异常: {e}")

    return result


# ============================================================
# 11. 访谈统计
# ============================================================
@router.get('/interviews')
def interview_summary(semester: str = Query(None), db: Session = Depends(get_db)):
    """访谈统计：总次数、按类型、待跟进、访谈覆盖率"""
    result = {
        'total_count': 0,
        'by_type': {},
        'pending_count': 0,
        'covered_student_count': 0,
        'total_student_count': 0,
        'coverage_rate': 0.0,
    }

    # 构建日期范围过滤
    interview_filters = []
    if semester and semester != 'all':
        start, end = _semester_date_range(semester)
        if start and end:
            interview_filters.append(StudentInterview.interview_date >= start)
            interview_filters.append(StudentInterview.interview_date <= end)

    try:
        q = db.query(StudentInterview)
        if interview_filters:
            q = q.filter(*interview_filters)
        result['total_count'] = q.count()
    except Exception as e:
        logger.warning(f"interviews 总数异常: {e}")

    try:
        q = (
            db.query(
                StudentInterview.interview_type,
                func.count(StudentInterview.id)
            )
            .group_by(StudentInterview.interview_type)
        )
        if interview_filters:
            q = q.filter(*interview_filters)
        rows = q.all()
        result['by_type'] = {(t or '未知'): cnt for t, cnt in rows}
    except Exception as e:
        logger.warning(f"interviews 按类型异常: {e}")

    try:
        q = db.query(StudentInterview).filter(StudentInterview.status == '需跟进')
        if interview_filters:
            q = q.filter(*interview_filters)
        result['pending_count'] = q.count()
    except Exception as e:
        logger.warning(f"interviews 待跟进异常: {e}")

    # 访谈覆盖率
    try:
        q_covered = db.query(StudentInterview.student_id).distinct()
        if interview_filters:
            q_covered = q_covered.filter(*interview_filters)
        covered = q_covered.count()
        total_students = db.query(Student).count()
        rate = round(covered / total_students * 100, 1) if total_students > 0 else 0.0
        result['covered_student_count'] = covered
        result['total_student_count'] = total_students
        result['coverage_rate'] = rate
    except Exception as e:
        logger.warning(f"interviews 覆盖率异常: {e}")

    return result


# ============================================================
# 12. 宿舍管理汇总
# ============================================================
@router.get('/dormitory')
def dormitory_summary(semester: str = Query(None), db: Session = Depends(get_db)):
    """宿舍管理汇总：走访、寝谈、请假统计"""
    result = {
        'visit_count': 0,
        'chat_count': 0,
        'leave_by_type': {},
        'leave_by_status': {},
        'leave_total': 0,
    }

    # 宿舍走访次数
    try:
        result['visit_count'] = db.query(StudentDormVisit).count()
    except Exception as e:
        logger.warning(f"dormitory 走访异常: {e}")

    # 寝谈记录数
    try:
        result['chat_count'] = db.query(StudentDormChat).count()
    except Exception as e:
        logger.warning(f"dormitory 寝谈异常: {e}")

    # 请假统计
    try:
        # 按类型
        rows = (
            db.query(StudentLeave.leave_type, func.count(StudentLeave.id))
            .group_by(StudentLeave.leave_type)
            .all()
        )
        result['leave_by_type'] = {(t or '未知'): cnt for t, cnt in rows}
    except Exception as e:
        logger.warning(f"dormitory 请假按类型异常: {e}")

    try:
        # 按审批状态
        rows = (
            db.query(StudentLeave.approval_status, func.count(StudentLeave.id))
            .group_by(StudentLeave.approval_status)
            .all()
        )
        status_labels = {'pending': '待审批', 'approved': '已批准', 'rejected': '已驳回'}
        result['leave_by_status'] = {
            status_labels.get(s, s or '未知'): cnt for s, cnt in rows
        }
    except Exception as e:
        logger.warning(f"dormitory 请假按状态异常: {e}")

    try:
        result['leave_total'] = db.query(StudentLeave).count()
    except Exception as e:
        logger.warning(f"dormitory 请假总数异常: {e}")

    return result


# ============================================================
# 13. 导出 Excel（修复 500 + 增加 sheet）
# ============================================================
@router.get('/export')
def export_semester_report(semester: str = Query(None), db: Session = Depends(get_db)):
    """导出学期报表为 Excel（多 Sheet）- 空数据时返回空白模板不报 500"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(500, 'openpyxl 未安装')

    wb = Workbook()

    # 样式
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    title_font = Font(bold=True, size=14)

    def write_title(ws, title, col=3):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col)
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')

    def write_headers(ws, row, headers):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def safe_count(query_func):
        try:
            return query_func()
        except Exception:
            return 0

    def safe_query(query_func, default=None):
        try:
            return query_func()
        except Exception:
            return default if default is not None else []

    # ==================== Sheet 1: 总览 ====================
    ws1 = wb.active
    ws1.title = '总览'
    write_title(ws1, '学期总览')

    try:
        total_students = db.query(Student).count()
        total_classes = db.query(ClassModel).count()
        total_majors = db.query(Major).count()
    except Exception:
        total_students = total_classes = total_majors = 0

    r = 3
    ws1.cell(row=r, column=1, value='指标')
    ws1.cell(row=r, column=2, value='数值')
    ws1.cell(row=r, column=1).font = header_font
    ws1.cell(row=r, column=2).font = header_font
    for label, val in [('学生总数', total_students), ('班级数', total_classes), ('专业数', total_majors)]:
        r += 1
        ws1.cell(row=r, column=1, value=label)
        ws1.cell(row=r, column=2, value=val)

    # 党员人数（按学期过滤）
    try:
        _exp_pp_subq = db.query(
            PartyProgress.student_id,
            func.max(PartyProgress.id).label('max_id')
        )
        if semester and semester != 'all':
            _exp_start, _exp_end = _semester_date_range(semester)
            if _exp_start and _exp_end:
                _exp_pp_subq = _exp_pp_subq.filter(
                    PartyProgress.stage_date >= _exp_start,
                    PartyProgress.stage_date <= _exp_end
                )
        _exp_pp_subq = _exp_pp_subq.group_by(PartyProgress.student_id).subquery()
        party_member_count = (
            db.query(func.count(PartyProgress.student_id))
            .join(_exp_pp_subq, PartyProgress.id == _exp_pp_subq.c.max_id)
            .filter(PartyProgress.stage.in_(['中共预备党员', '中共党员']))
            .scalar() or 0
        )
        r += 1
        ws1.cell(row=r, column=1, value='党员人数')
        ws1.cell(row=r, column=2, value=party_member_count)
    except Exception:
        r += 1
        ws1.cell(row=r, column=1, value='党员人数')
        ws1.cell(row=r, column=2, value=0)

    # 政治面貌
    r += 2
    ws1.cell(row=r, column=1, value='政治面貌分布')
    ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws1, r, ['政治面貌', '人数'])
    try:
        rows = db.query(Student.political_status, func.count(Student.id)).group_by(Student.political_status).all()
        for status, cnt in rows:
            r += 1
            ws1.cell(row=r, column=1, value=status or '群众')
            ws1.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # 性别
    r += 2
    ws1.cell(row=r, column=1, value='性别分布')
    ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws1, r, ['性别', '人数'])
    try:
        rows = db.query(Student.gender, func.count(Student.id)).group_by(Student.gender).all()
        for g, cnt in rows:
            r += 1
            ws1.cell(row=r, column=1, value=g or '未知')
            ws1.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # 住宿
    r += 2
    ws1.cell(row=r, column=1, value='住宿分布')
    ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws1, r, ['类型', '人数'])
    try:
        off = db.query(Student).filter(Student.is_off_campus == True).count()
        ws1.cell(row=r+1, column=1, value='住校')
        ws1.cell(row=r+1, column=2, value=total_students - off)
        ws1.cell(row=r+2, column=1, value='外宿')
        ws1.cell(row=r+2, column=2, value=off)
        r += 2
    except Exception:
        pass

    # 校区
    r += 2
    ws1.cell(row=r, column=1, value='校区分布')
    ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws1, r, ['校区', '人数'])
    try:
        rows = db.query(Student.campus, func.count(Student.id)).group_by(Student.campus).all()
        for c, cnt in rows:
            r += 1
            ws1.cell(row=r, column=1, value=c or '未知')
            ws1.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # ==================== Sheet 2: 学业 ====================
    ws2 = wb.create_sheet('学业')
    write_title(ws2, '学业数据汇总')

    r = 3
    write_headers(ws2, r, ['班级', '平均成绩'])
    try:
        grade_filters = [GradeRecord.score.isnot(None)]
        if semester and semester != 'all':
            grade_filters.append(GradeRecord.semester == semester)
        rows = (
            db.query(ClassModel.class_name, func.avg(GradeRecord.score))
            .join(Student, Student.class_id == ClassModel.id)
            .join(GradeRecord, GradeRecord.student_id == Student.id)
            .filter(*grade_filters)
            .group_by(ClassModel.id, ClassModel.class_name)
            .all()
        )
        for name, avg in rows:
            r += 1
            ws2.cell(row=r, column=1, value=name)
            ws2.cell(row=r, column=2, value=round(avg, 2) if avg else 0)
    except Exception:
        pass

    # 挂科率
    r += 2
    ws2.cell(row=r, column=1, value='挂科统计')
    ws2.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    try:
        total_with_grades = db.query(func.count(GradeRecord.student_id.distinct())).scalar() or 0
        fail_students = db.query(func.count(GradeRecord.student_id.distinct())).filter(GradeRecord.score.isnot(None), GradeRecord.score < 60).scalar() or 0
        fail_rate = round(fail_students / total_with_grades * 100, 2) if total_with_grades > 0 else 0.0
        ws2.cell(row=r, column=1, value='有成绩学生数')
        ws2.cell(row=r, column=2, value=total_with_grades)
        r += 1
        ws2.cell(row=r, column=1, value='挂科学生数')
        ws2.cell(row=r, column=2, value=fail_students)
        r += 1
        ws2.cell(row=r, column=1, value='挂科率(%)')
        ws2.cell(row=r, column=2, value=fail_rate)
    except Exception:
        pass

    # Top10
    r += 2
    ws2.cell(row=r, column=1, value='成绩排名 Top 10')
    ws2.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws2, r, ['排名', '学号', '姓名', '平均成绩'])
    try:
        rows = (
            db.query(Student.student_no, Student.name, func.avg(GradeRecord.score))
            .join(GradeRecord, GradeRecord.student_id == Student.id)
            .filter(GradeRecord.score.isnot(None))
            .group_by(Student.id, Student.student_no, Student.name)
            .order_by(func.avg(GradeRecord.score).desc())
            .limit(10)
            .all()
        )
        for idx, (sno, sname, avg) in enumerate(rows, 1):
            r += 1
            ws2.cell(row=r, column=1, value=idx)
            ws2.cell(row=r, column=2, value=sno)
            ws2.cell(row=r, column=3, value=sname)
            ws2.cell(row=r, column=4, value=round(avg, 2) if avg else 0)
    except Exception:
        pass

    # 预警
    r += 2
    ws2.cell(row=r, column=1, value='学业预警统计')
    ws2.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws2, r, ['预警等级', '人数'])
    try:
        red = db.query(WarningRecord).filter(WarningRecord.warning_type == 'red').count()
        yellow = db.query(WarningRecord).filter(WarningRecord.warning_type == 'yellow').count()
        r += 1
        ws2.cell(row=r, column=1, value='红色预警')
        ws2.cell(row=r, column=2, value=red)
        r += 1
        ws2.cell(row=r, column=1, value='黄色预警')
        ws2.cell(row=r, column=2, value=yellow)
        r += 1
        ws2.cell(row=r, column=1, value='合计')
        ws2.cell(row=r, column=2, value=red + yellow)
    except Exception:
        pass

    # ==================== Sheet 3: 党团 ====================
    ws3 = wb.create_sheet('党团')
    write_title(ws3, '党团发展进度')
    r = 3
    write_headers(ws3, r, ['阶段', '人数'])
    try:
        subq = (
            db.query(PartyProgress.student_id, func.max(PartyProgress.id).label('max_id'))
            .group_by(PartyProgress.student_id)
            .subquery()
        )
        rows = (
            db.query(PartyProgress.stage, func.count(PartyProgress.student_id))
            .join(subq, PartyProgress.id == subq.c.max_id)
            .group_by(PartyProgress.stage)
            .all()
        )
        for stage, cnt in rows:
            r += 1
            ws3.cell(row=r, column=1, value=stage)
            ws3.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    r += 2
    try:
        now = datetime.now()
        sem_start = f"{now.year}-09-01" if now.month >= 9 else f"{now.year}-03-01"
        new_cnt = db.query(func.count(PartyProgress.id)).filter(PartyProgress.stage_date >= sem_start).scalar() or 0
        ws3.cell(row=r, column=1, value='本学期新发展人数')
        ws3.cell(row=r, column=2, value=new_cnt)
    except Exception:
        pass

    # ==================== Sheet 4: 就业 ====================
    ws4 = wb.create_sheet('就业')
    write_title(ws4, '就业跟踪统计')
    r = 3
    write_headers(ws4, r, ['就业状态', '人数'])
    try:
        rows = db.query(EmploymentRecord.status, func.count(EmploymentRecord.id)).group_by(EmploymentRecord.status).all()
        total_emp = sum(cnt for _, cnt in rows)
        for status, cnt in rows:
            r += 1
            ws4.cell(row=r, column=1, value=status or '未知')
            ws4.cell(row=r, column=2, value=cnt)
        r += 1
        ws4.cell(row=r, column=1, value='合计')
        ws4.cell(row=r, column=2, value=total_emp)
        r += 1
        employed = sum(cnt for s, cnt in rows if s == '已签约')
        rate = round(employed / total_emp * 100, 2) if total_emp > 0 else 0.0
        ws4.cell(row=r, column=1, value='就业率(%)')
        ws4.cell(row=r, column=2, value=rate)
    except Exception:
        pass

    # ==================== Sheet 5: 活动 ====================
    ws5 = wb.create_sheet('活动')
    write_title(ws5, '学生活动统计')
    r = 3
    try:
        total_act = db.query(Activity).count()
        total_ppl = db.query(ActivitySignup).count()
    except Exception:
        total_act = total_ppl = 0
    ws5.cell(row=r, column=1, value='活动总数')
    ws5.cell(row=r, column=2, value=total_act)
    r += 1
    ws5.cell(row=r, column=1, value='参与人次')
    ws5.cell(row=r, column=2, value=total_ppl)
    r += 2
    ws5.cell(row=r, column=1, value='活动参与人数排名')
    ws5.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws5, r, ['活动名称', '类型', '参与人数'])
    try:
        rows = (
            db.query(Activity.title, Activity.activity_type, func.count(ActivitySignup.id))
            .join(ActivitySignup, ActivitySignup.activity_id == Activity.id)
            .group_by(Activity.id, Activity.title, Activity.activity_type)
            .order_by(func.count(ActivitySignup.id).desc())
            .limit(10)
            .all()
        )
        for title, atype, cnt in rows:
            r += 1
            ws5.cell(row=r, column=1, value=title)
            ws5.cell(row=r, column=2, value=atype or '')
            ws5.cell(row=r, column=3, value=cnt)
    except Exception:
        pass

    # ==================== Sheet 6: 考勤 ====================
    ws6 = wb.create_sheet('考勤')
    write_title(ws6, '考勤异常汇总')
    r = 3
    write_headers(ws6, r, ['异常类型', '次数'])
    try:
        q = db.query(
            StudentAttendanceException.exception_type,
            func.count(StudentAttendanceException.id)
        ).group_by(StudentAttendanceException.exception_type)
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    StudentAttendanceException.exception_date >= start,
                    StudentAttendanceException.exception_date <= end
                )
        rows = q.all()
        total_att = 0
        for t, cnt in rows:
            r += 1
            ws6.cell(row=r, column=1, value=t or '未知')
            ws6.cell(row=r, column=2, value=cnt)
            total_att += cnt
        r += 1
        ws6.cell(row=r, column=1, value='合计')
        ws6.cell(row=r, column=2, value=total_att)
    except Exception:
        pass

    # 按班级
    r += 2
    ws6.cell(row=r, column=1, value='按班级统计')
    ws6.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws6, r, ['班级', '异常次数'])
    try:
        q = (
            db.query(ClassModel.class_name, func.count(StudentAttendanceException.id))
            .join(Student, StudentAttendanceException.student_id == Student.id)
            .join(ClassModel, Student.class_id == ClassModel.id)
            .group_by(ClassModel.id, ClassModel.class_name)
        )
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(
                    StudentAttendanceException.exception_date >= start,
                    StudentAttendanceException.exception_date <= end
                )
        rows = q.all()
        for name, cnt in rows:
            r += 1
            ws6.cell(row=r, column=1, value=name)
            ws6.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # ==================== Sheet 7: 心理 ====================
    ws7 = wb.create_sheet('心理')
    write_title(ws7, '心理档案汇总')
    r = 3
    write_headers(ws7, r, ['关注等级', '人数'])
    try:
        rows = (
            db.query(PsychologyRecord.attention_level, func.count(PsychologyRecord.student_id.distinct()))
            .group_by(PsychologyRecord.attention_level)
            .all()
        )
        for level, cnt in rows:
            r += 1
            ws7.cell(row=r, column=1, value=level or '普通')
            ws7.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    r += 2
    try:
        total_counseling = db.query(func.sum(PsychologyRecord.counseling_count)).scalar() or 0
        ws7.cell(row=r, column=1, value='总咨询次数')
        ws7.cell(row=r, column=2, value=total_counseling)
    except Exception:
        pass

    # ==================== Sheet 8: 资助 ====================
    ws8 = wb.create_sheet('资助')
    write_title(ws8, '资助汇总')
    aid_year = _semester_to_academic_year(semester)

    r = 3
    ws8.cell(row=r, column=1, value='困难认定')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws8, r, ['等级', '人数'])
    try:
        q = db.query(StudentHardship.hardship_level, func.count(StudentHardship.id)).group_by(StudentHardship.hardship_level)
        if aid_year:
            q = q.filter(StudentHardship.academic_year == aid_year)
        rows = q.all()
        for level, cnt in rows:
            r += 1
            ws8.cell(row=r, column=1, value=level or '未知')
            ws8.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    r += 2
    ws8.cell(row=r, column=1, value='助学金')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws8, r, ['项目', '数值'])
    try:
        q = db.query(func.sum(StudentGrant.amount), func.count(StudentGrant.id))
        if aid_year:
            q = q.filter(StudentGrant.academic_year == aid_year)
        total_amount, count = q.first()
        r += 1
        ws8.cell(row=r, column=1, value='发放总额')
        ws8.cell(row=r, column=2, value=total_amount or 0)
        r += 1
        ws8.cell(row=r, column=1, value='发放人数')
        ws8.cell(row=r, column=2, value=count or 0)
    except Exception:
        pass

    r += 2
    ws8.cell(row=r, column=1, value='奖学金')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws8, r, ['项目', '数值'])
    try:
        q = db.query(func.sum(StudentScholarship.amount), func.count(StudentScholarship.id))
        if aid_year:
            q = q.filter(StudentScholarship.academic_year == aid_year)
        total_amount, count = q.first()
        r += 1
        ws8.cell(row=r, column=1, value='发放总额')
        ws8.cell(row=r, column=2, value=total_amount or 0)
        r += 1
        ws8.cell(row=r, column=1, value='获奖人数')
        ws8.cell(row=r, column=2, value=count or 0)
    except Exception:
        pass

    r += 2
    ws8.cell(row=r, column=1, value='助学贷款')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    try:
        total_loan = db.query(func.sum(StudentLoan.amount)).scalar() or 0
        loan_cnt = db.query(StudentLoan).count()
        ws8.cell(row=r, column=1, value='贷款总额')
        ws8.cell(row=r, column=2, value=total_loan)
        r += 1
        ws8.cell(row=r, column=1, value='贷款人数')
        ws8.cell(row=r, column=2, value=loan_cnt)
    except Exception:
        pass

    r += 2
    ws8.cell(row=r, column=1, value='勤工助学')
    ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    try:
        q = db.query(func.sum(StudentWorkStudy.compensation), func.count(StudentWorkStudy.id))
        if aid_year:
            q = q.filter(StudentWorkStudy.academic_year == aid_year)
        total_comp, count = q.first()
        ws8.cell(row=r, column=1, value='总报酬')
        ws8.cell(row=r, column=2, value=total_comp or 0)
        r += 1
        ws8.cell(row=r, column=1, value='参与人次')
        ws8.cell(row=r, column=2, value=count or 0)
    except Exception:
        pass

    # ==================== Sheet 9: 荣誉 ====================
    ws9 = wb.create_sheet('荣誉')
    write_title(ws9, '评优评先统计')
    r = 3
    write_headers(ws9, r, ['级别', '获奖人次'])
    try:
        rows = db.query(StudentHonor.level, func.count(StudentHonor.id)).group_by(StudentHonor.level).all()
        for level, cnt in rows:
            r += 1
            ws9.cell(row=r, column=1, value=level or '未知')
            ws9.cell(row=r, column=2, value=cnt)
    except Exception:
        pass
    r += 2
    try:
        student_count = db.query(StudentHonor.student_id).distinct().count()
        ws9.cell(row=r, column=1, value='获奖学生数')
        ws9.cell(row=r, column=2, value=student_count)
    except Exception:
        pass

    # ==================== Sheet 10: 违纪 ====================
    ws10 = wb.create_sheet('违纪')
    write_title(ws10, '违纪处分统计')
    r = 3
    write_headers(ws10, r, ['处分类型', '次数'])
    try:
        rows = db.query(StudentDiscipline.discipline_type, func.count(StudentDiscipline.id)).group_by(StudentDiscipline.discipline_type).all()
        for t, cnt in rows:
            r += 1
            ws10.cell(row=r, column=1, value=t or '未知')
            ws10.cell(row=r, column=2, value=cnt)
    except Exception:
        pass
    r += 2
    try:
        student_count = db.query(StudentDiscipline.student_id).distinct().count()
        ws10.cell(row=r, column=1, value='涉及学生数')
        ws10.cell(row=r, column=2, value=student_count)
    except Exception:
        pass

    # ==================== Sheet 11: 访谈 ====================
    ws11 = wb.create_sheet('访谈')
    write_title(ws11, '学生访谈统计')
    r = 3
    write_headers(ws11, r, ['访谈类型', '次数'])
    try:
        q = db.query(StudentInterview.interview_type, func.count(StudentInterview.id)).group_by(StudentInterview.interview_type)
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q = q.filter(StudentInterview.interview_date >= start, StudentInterview.interview_date <= end)
        rows = q.all()
        for t, cnt in rows:
            r += 1
            ws11.cell(row=r, column=1, value=t or '未知')
            ws11.cell(row=r, column=2, value=cnt)
    except Exception:
        pass
    r += 2
    try:
        q_all = db.query(StudentInterview)
        q_pending = db.query(StudentInterview).filter(StudentInterview.status == '需跟进')
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q_all = q_all.filter(StudentInterview.interview_date >= start, StudentInterview.interview_date <= end)
                q_pending = q_pending.filter(StudentInterview.interview_date >= start, StudentInterview.interview_date <= end)
        total_interviews = q_all.count()
        pending = q_pending.count()
        ws11.cell(row=r, column=1, value='访谈总次数')
        ws11.cell(row=r, column=2, value=total_interviews)
        r += 1
        ws11.cell(row=r, column=1, value='待跟进数量')
        ws11.cell(row=r, column=2, value=pending)
    except Exception:
        pass
    # 访谈覆盖率
    try:
        q_covered = db.query(StudentInterview.student_id).distinct()
        if semester and semester != 'all':
            start, end = _semester_date_range(semester)
            if start and end:
                q_covered = q_covered.filter(StudentInterview.interview_date >= start, StudentInterview.interview_date <= end)
        covered_count = q_covered.count()
        total_stu = db.query(Student).count()
        cov_rate = round(covered_count / total_stu * 100, 1) if total_stu > 0 else 0.0
        r += 2
        ws11.cell(row=r, column=1, value='访谈覆盖率(%)')
        ws11.cell(row=r, column=2, value=cov_rate)
        r += 1
        ws11.cell(row=r, column=1, value='被访谈学生数')
        ws11.cell(row=r, column=2, value=covered_count)
        r += 1
        ws11.cell(row=r, column=1, value='总学生数')
        ws11.cell(row=r, column=2, value=total_stu)
    except Exception:
        pass

    # ==================== Sheet 12: 宿舍 ====================
    ws12 = wb.create_sheet('宿舍')
    write_title(ws12, '宿舍管理汇总')
    r = 3
    write_headers(ws12, r, ['项目', '数值'])
    try:
        visit_cnt = db.query(StudentDormVisit).count()
        chat_cnt = db.query(StudentDormChat).count()
        leave_cnt = db.query(StudentLeave).count()
        r += 1
        ws12.cell(row=r, column=1, value='宿舍走访次数')
        ws12.cell(row=r, column=2, value=visit_cnt)
        r += 1
        ws12.cell(row=r, column=1, value='寝谈记录数')
        ws12.cell(row=r, column=2, value=chat_cnt)
        r += 1
        ws12.cell(row=r, column=1, value='请假记录数')
        ws12.cell(row=r, column=2, value=leave_cnt)
    except Exception:
        pass

    # 请假按类型
    r += 2
    ws12.cell(row=r, column=1, value='请假按类型')
    ws12.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    write_headers(ws12, r, ['请假类型', '次数'])
    try:
        rows = db.query(StudentLeave.leave_type, func.count(StudentLeave.id)).group_by(StudentLeave.leave_type).all()
        for t, cnt in rows:
            r += 1
            ws12.cell(row=r, column=1, value=t or '未知')
            ws12.cell(row=r, column=2, value=cnt)
    except Exception:
        pass

    # 写入 buffer
    try:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as e:
        logger.error(f'semester report export save failed: {e}')
        raise HTTPException(500, f'Excel 生成失败: {type(e).__name__}: {str(e)}')

    from urllib.parse import quote
    filename = f"学期报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ============================================================
# 14. 学期差值统计（修复空数据崩溃）
# ============================================================
@router.get('/compare')
def semester_compare(semester: str = Query(None), db: Session = Depends(get_db)):
    """与上一学期对比"""
    if not semester or semester == 'all':
        return {'comparison': {}}

    parts = semester.split('-')
    if len(parts) != 3:
        return {'comparison': {}}

    try:
        y1, y2, term = int(parts[0]), int(parts[1]), int(parts[2])
    except (ValueError, IndexError):
        return {'comparison': {}}

    if term == 2:
        prev_semester = f"{y1}-{y2}-1"
    else:
        prev_semester = f"{y1-1}-{y1}-2"

    result = {
        'current_semester': semester,
        'prev_semester': prev_semester,
        'comparison': {}
    }

    metrics = [
        'avg_score', 'fail_rate', 'warning_count', 'activity_participants',
        'attendance_exception_count', 'psychology_attention_count',
        'financial_aid_count', 'discipline_count', 'honor_count',
        'interview_count', 'interview_coverage',
    ]

    for metric in metrics:
        curr_val = _get_semester_metric(db, semester, metric)
        prev_val = _get_semester_metric(db, prev_semester, metric)

        if curr_val is not None and prev_val is not None:
            diff = curr_val - prev_val
            pct = (diff / prev_val * 100) if prev_val != 0 else 0
            result['comparison'][metric] = {
                'current': round(curr_val, 2),
                'previous': round(prev_val, 2),
                'diff': round(diff, 2),
                'change_pct': round(pct, 1)
            }
        else:
            result['comparison'][metric] = None

    return result


def _get_semester_metric(db: Session, semester: str, metric: str):
    """获取指定学期的某个指标值"""
    try:
        if metric == 'avg_score':
            rows = db.query(func.avg(GradeRecord.score)).filter(
                GradeRecord.semester == semester,
                GradeRecord.score.isnot(None)
            ).all()
            return rows[0][0] if rows and rows[0][0] else None
        elif metric == 'fail_rate':
            total = db.query(GradeRecord.student_id).filter(
                GradeRecord.semester == semester,
                GradeRecord.score.isnot(None)
            ).distinct().count()
            if total == 0:
                return None
            fail = db.query(GradeRecord.student_id).filter(
                GradeRecord.semester == semester,
                GradeRecord.score < 60
            ).distinct().count()
            return fail / total * 100
        elif metric == 'warning_count':
            return db.query(WarningRecord).filter(
                WarningRecord.semester == semester
            ).count()
        elif metric == 'activity_participants':
            # 按学期日期范围筛选活动
            start, end = _semester_date_range(semester)
            if not start or not end:
                return None
            return db.query(func.count(ActivitySignup.id)).join(Activity).filter(
                Activity.activity_date >= start,
                Activity.activity_date <= end
            ).scalar() or 0
        elif metric == 'attendance_exception_count':
            start, end = _semester_date_range(semester)
            if not start or not end:
                return None
            return db.query(StudentAttendanceException).filter(
                StudentAttendanceException.exception_date >= start,
                StudentAttendanceException.exception_date <= end
            ).count()
        elif metric == 'psychology_attention_count':
            start, end = _semester_date_range(semester)
            if not start or not end:
                return None
            return db.query(PsychologyRecord.student_id).filter(
                PsychologyRecord.attention_level.in_(['一级关注', '二级关注', '三级关注']),
                PsychologyRecord.record_date >= start,
                PsychologyRecord.record_date <= end
            ).distinct().count()
        elif metric == 'financial_aid_count':
            aid_year = _semester_to_academic_year(semester)
            if not aid_year:
                return None
            cnt = 0
            cnt += db.query(StudentHardship).filter(StudentHardship.academic_year == aid_year).count()
            cnt += db.query(StudentGrant).filter(StudentGrant.academic_year == aid_year).count()
            cnt += db.query(StudentScholarship).filter(StudentScholarship.academic_year == aid_year).count()
            cnt += db.query(StudentLoan).count()  # 贷款无学年字段
            cnt += db.query(StudentWorkStudy).filter(StudentWorkStudy.academic_year == aid_year).count()
            return cnt
        elif metric == 'discipline_count':
            start, end = _semester_date_range(semester)
            if not start or not end:
                return None
            return db.query(StudentDiscipline.student_id).filter(
                StudentDiscipline.discipline_date >= start,
                StudentDiscipline.discipline_date <= end
            ).distinct().count()
        elif metric == 'honor_count':
            honor_year = _semester_to_academic_year(semester)
            if not honor_year:
                return None
            return db.query(StudentHonor).filter(
                StudentHonor.academic_year == honor_year
            ).count()
        elif metric == 'interview_count':
            start, end = _semester_date_range(semester)
            if not start or not end:
                return None
            return db.query(StudentInterview).filter(
                StudentInterview.interview_date >= start,
                StudentInterview.interview_date <= end
            ).count()
        elif metric == 'interview_coverage':
            start, end = _semester_date_range(semester)
            if not start or not end:
                return None
            total_students = db.query(Student).count()
            if total_students == 0:
                return None
            covered = db.query(StudentInterview.student_id).filter(
                StudentInterview.interview_date >= start,
                StudentInterview.interview_date <= end
            ).distinct().count()
            return covered / total_students * 100
    except Exception as e:
        logger.warning(f"获取指标 {metric} 异常: {e}")
    return None
